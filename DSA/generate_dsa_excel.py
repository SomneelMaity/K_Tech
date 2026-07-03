import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ─── Color Palette ────────────────────────────────────────────────────────────
HEADER_FILLS = {
    "easy":    PatternFill("solid", fgColor="4CAF50"),   # green
    "medium":  PatternFill("solid", fgColor="FF9800"),   # orange
    "hard":    PatternFill("solid", fgColor="F44336"),   # red
    "title":   PatternFill("solid", fgColor="1565C0"),   # dark blue
    "company": PatternFill("solid", fgColor="4A148C"),   # deep purple
    "plan":    PatternFill("solid", fgColor="00695C"),   # teal
    "col_hdr": PatternFill("solid", fgColor="1976D2"),   # blue
}
WHITE   = Font(color="FFFFFF", bold=True, size=11)
BOLD    = Font(bold=True, size=10)
NORMAL  = Font(size=10)
LINK    = Font(color="1565C0", underline="single", size=10)
CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN = Side(style="thin", color="BDBDBD")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DIFF_FILL = {
    "Easy":   PatternFill("solid", fgColor="C8E6C9"),
    "Medium": PatternFill("solid", fgColor="FFE0B2"),
    "Hard":   PatternFill("solid", fgColor="FFCDD2"),
}

# ─── Helper ───────────────────────────────────────────────────────────────────
def safe_title(name):
    """Strip characters invalid in Excel sheet names and truncate to 31 chars."""
    for ch in r"/\?*[]:'":
        name = name.replace(ch, "-")
    return name[:31]


def add_topic_sheet(wb, sheet_name, rows):
    """
    rows: list of (no, problem, difficulty, lc_number, url)
    """
    ws = wb.create_sheet(title=safe_title(sheet_name))

    # Title row
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = sheet_name
    c.fill  = HEADER_FILLS["title"]
    c.font  = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Column headers
    headers = ["#", "Problem", "Difficulty", "LC #", "LeetCode Link"]
    col_widths = [5, 48, 12, 8, 50]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill      = HEADER_FILLS["col_hdr"]
        cell.font      = WHITE
        cell.alignment = CENTER
        cell.border    = BOX
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    # Data rows
    for r_idx, (no, problem, difficulty, lc_num, url) in enumerate(rows, start=3):
        ws.cell(row=r_idx, column=1, value=no).alignment        = CENTER
        ws.cell(row=r_idx, column=2, value=problem).alignment   = LEFT
        ws.cell(row=r_idx, column=3, value=difficulty).alignment = CENTER
        ws.cell(row=r_idx, column=4, value=lc_num).alignment    = CENTER

        link_cell = ws.cell(row=r_idx, column=5,
                            value=f"LC {lc_num} — Open")
        link_cell.hyperlink  = url
        link_cell.font       = LINK
        link_cell.alignment  = LEFT

        fill = DIFF_FILL.get(difficulty, PatternFill())
        for col in range(1, 6):
            cell = ws.cell(row=r_idx, column=col)
            cell.fill   = fill
            cell.border = BOX
            if col != 5:
                cell.font = NORMAL
        ws.row_dimensions[r_idx].height = 18

    # Freeze panes below header
    ws.freeze_panes = "A3"
    return ws


def add_company_sheet(wb, company, rows):
    """
    rows: list of (problem, lc_number, url)
    """
    ws = wb.create_sheet(title=safe_title(company))

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"{company} — High-Frequency Questions"
    c.fill  = HEADER_FILLS["company"]
    c.font  = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    headers    = ["#", "Problem", "LC #", "LeetCode Link"]
    col_widths = [5, 55, 8, 52]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill      = HEADER_FILLS["col_hdr"]
        cell.font      = WHITE
        cell.alignment = CENTER
        cell.border    = BOX
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    ALT1 = PatternFill("solid", fgColor="EDE7F6")
    ALT2 = PatternFill("solid", fgColor="FFFFFF")

    for r_idx, (problem, lc_num, url) in enumerate(rows, start=3):
        fill = ALT1 if r_idx % 2 == 1 else ALT2
        ws.cell(row=r_idx, column=1, value=r_idx - 2).alignment = CENTER
        ws.cell(row=r_idx, column=2, value=problem).alignment   = LEFT
        ws.cell(row=r_idx, column=3, value=lc_num).alignment    = CENTER

        link_cell = ws.cell(row=r_idx, column=4,
                            value=f"LC {lc_num} — Open")
        link_cell.hyperlink  = url
        link_cell.font       = LINK
        link_cell.alignment  = LEFT

        for col in range(1, 5):
            cell = ws.cell(row=r_idx, column=col)
            cell.fill   = fill
            cell.border = BOX
            if col != 4:
                cell.font = NORMAL
        ws.row_dimensions[r_idx].height = 18

    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA
# ═══════════════════════════════════════════════════════════════════════════════

topics = {}

topics["Arrays & Hashing"] = [
    (1,  "Two Sum",                                  "Easy",   1,    "https://leetcode.com/problems/two-sum/"),
    (2,  "Contains Duplicate",                        "Easy",   217,  "https://leetcode.com/problems/contains-duplicate/"),
    (3,  "Valid Anagram",                             "Easy",   242,  "https://leetcode.com/problems/valid-anagram/"),
    (4,  "Group Anagrams",                            "Medium", 49,   "https://leetcode.com/problems/group-anagrams/"),
    (5,  "Top K Frequent Elements",                   "Medium", 347,  "https://leetcode.com/problems/top-k-frequent-elements/"),
    (6,  "Product of Array Except Self",              "Medium", 238,  "https://leetcode.com/problems/product-of-array-except-self/"),
    (7,  "Valid Sudoku",                              "Medium", 36,   "https://leetcode.com/problems/valid-sudoku/"),
    (8,  "Longest Consecutive Sequence",              "Medium", 128,  "https://leetcode.com/problems/longest-consecutive-sequence/"),
    (9,  "Subarray Sum Equals K",                     "Medium", 560,  "https://leetcode.com/problems/subarray-sum-equals-k/"),
    (10, "Sort Colors",                               "Medium", 75,   "https://leetcode.com/problems/sort-colors/"),
    (11, "Find All Duplicates in an Array",           "Medium", 442,  "https://leetcode.com/problems/find-all-duplicates-in-an-array/"),
    (12, "First Missing Positive",                    "Hard",   41,   "https://leetcode.com/problems/first-missing-positive/"),
    (13, "Maximum Points from Cards",                 "Medium", 1423, "https://leetcode.com/problems/maximum-points-you-can-obtain-from-cards/"),
]

topics["Two Pointers"] = [
    (1, "Valid Palindrome",                     "Easy",   125, "https://leetcode.com/problems/valid-palindrome/"),
    (2, "Two Sum II — Sorted Array",            "Medium", 167, "https://leetcode.com/problems/two-sum-ii-input-array-is-sorted/"),
    (3, "3Sum",                                 "Medium", 15,  "https://leetcode.com/problems/3sum/"),
    (4, "4Sum",                                 "Medium", 18,  "https://leetcode.com/problems/4sum/"),
    (5, "Container With Most Water",            "Medium", 11,  "https://leetcode.com/problems/container-with-most-water/"),
    (6, "Trapping Rain Water",                  "Hard",   42,  "https://leetcode.com/problems/trapping-rain-water/"),
    (7, "Move Zeroes",                          "Easy",   283, "https://leetcode.com/problems/move-zeroes/"),
    (8, "Remove Duplicates from Sorted Array",  "Easy",   26,  "https://leetcode.com/problems/remove-duplicates-from-sorted-array/"),
]

topics["Sliding Window"] = [
    (1, "Best Time to Buy and Sell Stock",             "Easy",   121, "https://leetcode.com/problems/best-time-to-buy-and-sell-stock/"),
    (2, "Longest Substring Without Repeating Chars",   "Medium", 3,   "https://leetcode.com/problems/longest-substring-without-repeating-characters/"),
    (3, "Longest Repeating Character Replacement",     "Medium", 424, "https://leetcode.com/problems/longest-repeating-character-replacement/"),
    (4, "Permutation in String",                       "Medium", 567, "https://leetcode.com/problems/permutation-in-string/"),
    (5, "Minimum Window Substring",                    "Hard",   76,  "https://leetcode.com/problems/minimum-window-substring/"),
    (6, "Sliding Window Maximum",                      "Hard",   239, "https://leetcode.com/problems/sliding-window-maximum/"),
    (7, "Find All Anagrams in a String",               "Medium", 438, "https://leetcode.com/problems/find-all-anagrams-in-a-string/"),
    (8, "Minimum Size Subarray Sum",                   "Medium", 209, "https://leetcode.com/problems/minimum-size-subarray-sum/"),
]

topics["Stack"] = [
    (1,  "Valid Parentheses",                      "Easy",   20,   "https://leetcode.com/problems/valid-parentheses/"),
    (2,  "Min Stack",                              "Medium", 155,  "https://leetcode.com/problems/min-stack/"),
    (3,  "Evaluate Reverse Polish Notation",       "Medium", 150,  "https://leetcode.com/problems/evaluate-reverse-polish-notation/"),
    (4,  "Generate Parentheses",                   "Medium", 22,   "https://leetcode.com/problems/generate-parentheses/"),
    (5,  "Daily Temperatures",                     "Medium", 739,  "https://leetcode.com/problems/daily-temperatures/"),
    (6,  "Car Fleet",                              "Medium", 853,  "https://leetcode.com/problems/car-fleet/"),
    (7,  "Largest Rectangle in Histogram",         "Hard",   84,   "https://leetcode.com/problems/largest-rectangle-in-histogram/"),
    (8,  "Next Greater Element I",                 "Easy",   496,  "https://leetcode.com/problems/next-greater-element-i/"),
    (9,  "Asteroid Collision",                     "Medium", 735,  "https://leetcode.com/problems/asteroid-collision/"),
    (10, "Minimum Remove to Make Valid Parens",    "Medium", 1249, "https://leetcode.com/problems/minimum-remove-to-make-valid-parentheses/"),
    (11, "Decode String",                          "Medium", 394,  "https://leetcode.com/problems/decode-string/"),
]

topics["Binary Search"] = [
    (1,  "Binary Search",                              "Easy",   704,  "https://leetcode.com/problems/binary-search/"),
    (2,  "Search a 2D Matrix",                         "Medium", 74,   "https://leetcode.com/problems/search-a-2d-matrix/"),
    (3,  "Koko Eating Bananas",                        "Medium", 875,  "https://leetcode.com/problems/koko-eating-bananas/"),
    (4,  "Find Minimum in Rotated Sorted Array",       "Medium", 153,  "https://leetcode.com/problems/find-minimum-in-rotated-sorted-array/"),
    (5,  "Search in Rotated Sorted Array",             "Medium", 33,   "https://leetcode.com/problems/search-in-rotated-sorted-array/"),
    (6,  "Time Based Key-Value Store",                 "Medium", 981,  "https://leetcode.com/problems/time-based-key-value-store/"),
    (7,  "Find First and Last Position of Element",    "Medium", 34,   "https://leetcode.com/problems/find-first-and-last-position-of-element-in-sorted-array/"),
    (8,  "Median of Two Sorted Arrays",               "Hard",   4,    "https://leetcode.com/problems/median-of-two-sorted-arrays/"),
    (9,  "Capacity to Ship Packages Within D Days",    "Medium", 1011, "https://leetcode.com/problems/capacity-to-ship-packages-within-d-days/"),
    (10, "Peak Index in a Mountain Array",             "Medium", 852,  "https://leetcode.com/problems/peak-index-in-a-mountain-array/"),
    (11, "Split Array Largest Sum",                    "Hard",   410,  "https://leetcode.com/problems/split-array-largest-sum/"),
]

topics["Linked List"] = [
    (1,  "Reverse Linked List",                   "Easy",   206, "https://leetcode.com/problems/reverse-linked-list/"),
    (2,  "Merge Two Sorted Lists",                "Easy",   21,  "https://leetcode.com/problems/merge-two-sorted-lists/"),
    (3,  "Linked List Cycle",                     "Easy",   141, "https://leetcode.com/problems/linked-list-cycle/"),
    (4,  "Reorder List",                          "Medium", 143, "https://leetcode.com/problems/reorder-list/"),
    (5,  "Remove Nth Node From End of List",      "Medium", 19,  "https://leetcode.com/problems/remove-nth-node-from-end-of-list/"),
    (6,  "Copy List with Random Pointer",         "Medium", 138, "https://leetcode.com/problems/copy-list-with-random-pointer/"),
    (7,  "Add Two Numbers",                       "Medium", 2,   "https://leetcode.com/problems/add-two-numbers/"),
    (8,  "Find the Duplicate Number",             "Medium", 287, "https://leetcode.com/problems/find-the-duplicate-number/"),
    (9,  "LRU Cache",                             "Medium", 146, "https://leetcode.com/problems/lru-cache/"),
    (10, "Merge K Sorted Lists",                  "Hard",   23,  "https://leetcode.com/problems/merge-k-sorted-lists/"),
    (11, "Reverse Nodes in K-Group",              "Hard",   25,  "https://leetcode.com/problems/reverse-nodes-in-k-group/"),
    (12, "Intersection of Two Linked Lists",      "Easy",   160, "https://leetcode.com/problems/intersection-of-two-linked-lists/"),
    (13, "Palindrome Linked List",                "Easy",   234, "https://leetcode.com/problems/palindrome-linked-list/"),
    (14, "Flatten a Multilevel Doubly LL",        "Medium", 430, "https://leetcode.com/problems/flatten-a-multilevel-doubly-linked-list/"),
]

topics["Trees"] = [
    (1,  "Invert Binary Tree",                        "Easy",   226, "https://leetcode.com/problems/invert-binary-tree/"),
    (2,  "Maximum Depth of Binary Tree",              "Easy",   104, "https://leetcode.com/problems/maximum-depth-of-binary-tree/"),
    (3,  "Diameter of Binary Tree",                   "Easy",   543, "https://leetcode.com/problems/diameter-of-binary-tree/"),
    (4,  "Balanced Binary Tree",                      "Easy",   110, "https://leetcode.com/problems/balanced-binary-tree/"),
    (5,  "Same Tree",                                 "Easy",   100, "https://leetcode.com/problems/same-tree/"),
    (6,  "Subtree of Another Tree",                   "Easy",   572, "https://leetcode.com/problems/subtree-of-another-tree/"),
    (7,  "Lowest Common Ancestor of BST",             "Medium", 235, "https://leetcode.com/problems/lowest-common-ancestor-of-a-binary-search-tree/"),
    (8,  "Lowest Common Ancestor of Binary Tree",     "Medium", 236, "https://leetcode.com/problems/lowest-common-ancestor-of-a-binary-tree/"),
    (9,  "Binary Tree Level Order Traversal",         "Medium", 102, "https://leetcode.com/problems/binary-tree-level-order-traversal/"),
    (10, "Binary Tree Right Side View",               "Medium", 199, "https://leetcode.com/problems/binary-tree-right-side-view/"),
    (11, "Count Good Nodes in Binary Tree",           "Medium", 1448,"https://leetcode.com/problems/count-good-nodes-in-binary-tree/"),
    (12, "Validate Binary Search Tree",               "Medium", 98,  "https://leetcode.com/problems/validate-binary-search-tree/"),
    (13, "Kth Smallest Element in a BST",             "Medium", 230, "https://leetcode.com/problems/kth-smallest-element-in-a-bst/"),
    (14, "Construct BT from Preorder & Inorder",      "Medium", 105, "https://leetcode.com/problems/construct-binary-tree-from-preorder-and-inorder-traversal/"),
    (15, "Binary Tree Maximum Path Sum",              "Hard",   124, "https://leetcode.com/problems/binary-tree-maximum-path-sum/"),
    (16, "Serialize and Deserialize Binary Tree",     "Hard",   297, "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/"),
    (17, "Path Sum II",                               "Medium", 113, "https://leetcode.com/problems/path-sum-ii/"),
    (18, "Flatten Binary Tree to Linked List",        "Medium", 114, "https://leetcode.com/problems/flatten-binary-tree-to-linked-list/"),
    (19, "Populating Next Right Pointers",            "Medium", 116, "https://leetcode.com/problems/populating-next-right-pointers-in-each-node/"),
    (20, "Recover Binary Search Tree",                "Medium", 99,  "https://leetcode.com/problems/recover-binary-search-tree/"),
    (21, "Vertical Order Traversal of Binary Tree",   "Hard",   987, "https://leetcode.com/problems/vertical-order-traversal-of-a-binary-tree/"),
    (22, "All Nodes Distance K in Binary Tree",       "Medium", 863, "https://leetcode.com/problems/all-nodes-distance-k-in-binary-tree/"),
]

topics["Heap / Priority Queue"] = [
    (1,  "Kth Largest Element in a Stream",  "Easy",   703, "https://leetcode.com/problems/kth-largest-element-in-a-stream/"),
    (2,  "Last Stone Weight",                "Easy",   1046,"https://leetcode.com/problems/last-stone-weight/"),
    (3,  "K Closest Points to Origin",       "Medium", 973, "https://leetcode.com/problems/k-closest-points-to-origin/"),
    (4,  "Kth Largest Element in an Array",  "Medium", 215, "https://leetcode.com/problems/kth-largest-element-in-an-array/"),
    (5,  "Task Scheduler",                   "Medium", 621, "https://leetcode.com/problems/task-scheduler/"),
    (6,  "Design Twitter",                   "Medium", 355, "https://leetcode.com/problems/design-twitter/"),
    (7,  "Find Median from Data Stream",     "Hard",   295, "https://leetcode.com/problems/find-median-from-data-stream/"),
    (8,  "IPO (Maximize Capital)",           "Hard",   502, "https://leetcode.com/problems/ipo/"),
    (9,  "Ugly Number II",                   "Medium", 264, "https://leetcode.com/problems/ugly-number-ii/"),
    (10, "Top K Frequent Words",             "Medium", 692, "https://leetcode.com/problems/top-k-frequent-words/"),
]

topics["Backtracking"] = [
    (1,  "Subsets",                              "Medium", 78,  "https://leetcode.com/problems/subsets/"),
    (2,  "Combination Sum",                      "Medium", 39,  "https://leetcode.com/problems/combination-sum/"),
    (3,  "Permutations",                         "Medium", 46,  "https://leetcode.com/problems/permutations/"),
    (4,  "Subsets II",                           "Medium", 90,  "https://leetcode.com/problems/subsets-ii/"),
    (5,  "Combination Sum II",                   "Medium", 40,  "https://leetcode.com/problems/combination-sum-ii/"),
    (6,  "Word Search",                          "Medium", 79,  "https://leetcode.com/problems/word-search/"),
    (7,  "Palindrome Partitioning",              "Medium", 131, "https://leetcode.com/problems/palindrome-partitioning/"),
    (8,  "Letter Combinations of Phone Number",  "Medium", 17,  "https://leetcode.com/problems/letter-combinations-of-a-phone-number/"),
    (9,  "N-Queens",                             "Hard",   51,  "https://leetcode.com/problems/n-queens/"),
    (10, "Sudoku Solver",                        "Hard",   37,  "https://leetcode.com/problems/sudoku-solver/"),
    (11, "Expression Add Operators",             "Hard",   282, "https://leetcode.com/problems/expression-add-operators/"),
    (12, "Restore IP Addresses",                 "Medium", 93,  "https://leetcode.com/problems/restore-ip-addresses/"),
]

topics["Tries"] = [
    (1, "Implement Trie (Prefix Tree)",               "Medium", 208, "https://leetcode.com/problems/implement-trie-prefix-tree/"),
    (2, "Design Add and Search Words Structure",       "Medium", 211, "https://leetcode.com/problems/design-add-and-search-words-data-structure/"),
    (3, "Word Search II",                             "Hard",   212, "https://leetcode.com/problems/word-search-ii/"),
    (4, "Replace Words",                              "Medium", 648, "https://leetcode.com/problems/replace-words/"),
    (5, "Maximum XOR of Two Numbers in Array",        "Medium", 421, "https://leetcode.com/problems/maximum-xor-of-two-numbers-in-an-array/"),
]

topics["Graphs"] = [
    (1,  "Number of Islands",                           "Medium", 200,  "https://leetcode.com/problems/number-of-islands/"),
    (2,  "Clone Graph",                                 "Medium", 133,  "https://leetcode.com/problems/clone-graph/"),
    (3,  "Max Area of Island",                          "Medium", 695,  "https://leetcode.com/problems/max-area-of-island/"),
    (4,  "Pacific Atlantic Water Flow",                  "Medium", 417,  "https://leetcode.com/problems/pacific-atlantic-water-flow/"),
    (5,  "Surrounded Regions",                          "Medium", 130,  "https://leetcode.com/problems/surrounded-regions/"),
    (6,  "Rotting Oranges",                             "Medium", 994,  "https://leetcode.com/problems/rotting-oranges/"),
    (7,  "Course Schedule",                             "Medium", 207,  "https://leetcode.com/problems/course-schedule/"),
    (8,  "Course Schedule II",                          "Medium", 210,  "https://leetcode.com/problems/course-schedule-ii/"),
    (9,  "Redundant Connection",                        "Medium", 684,  "https://leetcode.com/problems/redundant-connection/"),
    (10, "Word Ladder",                                 "Hard",   127,  "https://leetcode.com/problems/word-ladder/"),
    (11, "Is Graph Bipartite?",                         "Medium", 785,  "https://leetcode.com/problems/is-graph-bipartite/"),
    (12, "Find Eventual Safe States",                   "Medium", 802,  "https://leetcode.com/problems/find-eventual-safe-states/"),
    (13, "Number of Enclaves",                          "Medium", 1020, "https://leetcode.com/problems/number-of-enclaves/"),
    (14, "Shortest Path in Binary Matrix",              "Medium", 1091, "https://leetcode.com/problems/shortest-path-in-binary-matrix/"),
    (15, "Min Vertices to Reach All Nodes",             "Medium", 1557, "https://leetcode.com/problems/minimum-number-of-vertices-to-reach-all-nodes/"),
    (16, "Critical Connections in a Network",           "Hard",   1192, "https://leetcode.com/problems/critical-connections-in-a-network/"),
    (17, "All Paths From Source to Target",             "Medium", 797,  "https://leetcode.com/problems/all-paths-from-source-to-target/"),
    (18, "Accounts Merge",                              "Medium", 721,  "https://leetcode.com/problems/accounts-merge/"),
    (19, "Number of Provinces",                         "Medium", 547,  "https://leetcode.com/problems/number-of-provinces/"),
    (20, "Evaluate Division",                           "Medium", 399,  "https://leetcode.com/problems/evaluate-division/"),
]

topics["Advanced Graphs"] = [
    (1, "Reconstruct Itinerary",                       "Hard",   332,  "https://leetcode.com/problems/reconstruct-itinerary/"),
    (2, "Min Cost to Connect All Points",              "Medium", 1584, "https://leetcode.com/problems/min-cost-to-connect-all-points/"),
    (3, "Network Delay Time",                          "Medium", 743,  "https://leetcode.com/problems/network-delay-time/"),
    (4, "Swim in Rising Water",                        "Hard",   778,  "https://leetcode.com/problems/swim-in-rising-water/"),
    (5, "Cheapest Flights Within K Stops",             "Medium", 787,  "https://leetcode.com/problems/cheapest-flights-within-k-stops/"),
    (6, "Path With Minimum Effort",                    "Medium", 1631, "https://leetcode.com/problems/path-with-minimum-effort/"),
    (7, "Find City With Smallest Neighbor Count",      "Medium", 1334, "https://leetcode.com/problems/find-the-city-with-the-smallest-number-of-neighbors-at-a-threshold-distance/"),
    (8, "Parallel Courses III",                        "Hard",   2050, "https://leetcode.com/problems/parallel-courses-iii/"),
]

topics["DP — 1D"] = [
    (1,  "Climbing Stairs",                    "Easy",   70,   "https://leetcode.com/problems/climbing-stairs/"),
    (2,  "Min Cost Climbing Stairs",           "Easy",   746,  "https://leetcode.com/problems/min-cost-climbing-stairs/"),
    (3,  "House Robber",                       "Medium", 198,  "https://leetcode.com/problems/house-robber/"),
    (4,  "House Robber II",                    "Medium", 213,  "https://leetcode.com/problems/house-robber-ii/"),
    (5,  "Longest Palindromic Substring",      "Medium", 5,    "https://leetcode.com/problems/longest-palindromic-substring/"),
    (6,  "Palindromic Substrings",             "Medium", 647,  "https://leetcode.com/problems/palindromic-substrings/"),
    (7,  "Decode Ways",                        "Medium", 91,   "https://leetcode.com/problems/decode-ways/"),
    (8,  "Coin Change",                        "Medium", 322,  "https://leetcode.com/problems/coin-change/"),
    (9,  "Maximum Product Subarray",           "Medium", 152,  "https://leetcode.com/problems/maximum-product-subarray/"),
    (10, "Word Break",                         "Medium", 139,  "https://leetcode.com/problems/word-break/"),
    (11, "Longest Increasing Subsequence",     "Medium", 300,  "https://leetcode.com/problems/longest-increasing-subsequence/"),
    (12, "Partition Equal Subset Sum",         "Medium", 416,  "https://leetcode.com/problems/partition-equal-subset-sum/"),
    (13, "Jump Game",                          "Medium", 55,   "https://leetcode.com/problems/jump-game/"),
    (14, "Jump Game II",                       "Medium", 45,   "https://leetcode.com/problems/jump-game-ii/"),
    (15, "Maximum Subarray (Kadane's)",        "Medium", 53,   "https://leetcode.com/problems/maximum-subarray/"),
    (16, "Perfect Squares",                    "Medium", 279,  "https://leetcode.com/problems/perfect-squares/"),
    (17, "Fibonacci Number",                   "Easy",   509,  "https://leetcode.com/problems/fibonacci-number/"),
    (18, "N-th Tribonacci Number",             "Easy",   1137, "https://leetcode.com/problems/n-th-tribonacci-number/"),
]

topics["DP — 2D"] = [
    (1,  "Unique Paths",                               "Medium", 62,   "https://leetcode.com/problems/unique-paths/"),
    (2,  "Unique Paths II",                            "Medium", 63,   "https://leetcode.com/problems/unique-paths-ii/"),
    (3,  "Minimum Path Sum",                           "Medium", 64,   "https://leetcode.com/problems/minimum-path-sum/"),
    (4,  "Longest Common Subsequence",                 "Medium", 1143, "https://leetcode.com/problems/longest-common-subsequence/"),
    (5,  "Best Time to Buy Sell — Cooldown",           "Medium", 309,  "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-with-cooldown/"),
    (6,  "Best Time to Buy Sell — Fee",                "Medium", 714,  "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-with-transaction-fee/"),
    (7,  "Coin Change II",                             "Medium", 518,  "https://leetcode.com/problems/coin-change-ii/"),
    (8,  "Target Sum",                                 "Medium", 494,  "https://leetcode.com/problems/target-sum/"),
    (9,  "Interleaving String",                        "Medium", 97,   "https://leetcode.com/problems/interleaving-string/"),
    (10, "Longest Increasing Path in Matrix",          "Hard",   329,  "https://leetcode.com/problems/longest-increasing-path-in-a-matrix/"),
    (11, "Distinct Subsequences",                      "Hard",   115,  "https://leetcode.com/problems/distinct-subsequences/"),
    (12, "Edit Distance",                              "Medium", 72,   "https://leetcode.com/problems/edit-distance/"),
    (13, "Burst Balloons",                             "Hard",   312,  "https://leetcode.com/problems/burst-balloons/"),
    (14, "Regular Expression Matching",                "Hard",   10,   "https://leetcode.com/problems/regular-expression-matching/"),
    (15, "Wildcard Matching",                          "Hard",   44,   "https://leetcode.com/problems/wildcard-matching/"),
    (16, "Maximum Rectangle",                          "Hard",   85,   "https://leetcode.com/problems/maximal-rectangle/"),
    (17, "Maximal Square",                             "Medium", 221,  "https://leetcode.com/problems/maximal-square/"),
    (18, "Triangle",                                   "Medium", 120,  "https://leetcode.com/problems/triangle/"),
    (19, "Best Time to Buy Sell Stock III",            "Hard",   123,  "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-iii/"),
    (20, "Best Time to Buy Sell Stock IV",             "Hard",   188,  "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-iv/"),
    (21, "Stone Game",                                 "Medium", 877,  "https://leetcode.com/problems/stone-game/"),
    (22, "Minimum Difficulty of a Job Schedule",       "Hard",   1335, "https://leetcode.com/problems/minimum-difficulty-of-a-job-schedule/"),
]

topics["Greedy"] = [
    (1, "Gas Station",                              "Medium", 134,  "https://leetcode.com/problems/gas-station/"),
    (2, "Hand of Straights",                        "Medium", 846,  "https://leetcode.com/problems/hand-of-straights/"),
    (3, "Partition Labels",                         "Medium", 763,  "https://leetcode.com/problems/partition-labels/"),
    (4, "Valid Parenthesis String",                 "Medium", 678,  "https://leetcode.com/problems/valid-parenthesis-string/"),
    (5, "Merge Triplets to Form Target Triplet",   "Medium", 1899, "https://leetcode.com/problems/merge-triplets-to-form-a-target-triplet/"),
    (6, "Boats to Save People",                    "Medium", 881,  "https://leetcode.com/problems/boats-to-save-people/"),
    (7, "Minimum Arrows to Burst Balloons",        "Medium", 452,  "https://leetcode.com/problems/minimum-number-of-arrows-to-burst-balloons/"),
    (8, "Candy",                                   "Hard",   135,  "https://leetcode.com/problems/candy/"),
    (9, "Assign Cookies",                          "Easy",   455,  "https://leetcode.com/problems/assign-cookies/"),
]

topics["Intervals"] = [
    (1, "Insert Interval",                       "Medium", 57,   "https://leetcode.com/problems/insert-interval/"),
    (2, "Merge Intervals",                       "Medium", 56,   "https://leetcode.com/problems/merge-intervals/"),
    (3, "Non-overlapping Intervals",             "Medium", 435,  "https://leetcode.com/problems/non-overlapping-intervals/"),
    (4, "Minimum Interval per Query",            "Hard",   1851, "https://leetcode.com/problems/minimum-interval-to-include-each-query/"),
    (5, "Employee Free Time",                    "Hard",   759,  "https://leetcode.com/problems/employee-free-time/"),
]

topics["Math & Geometry"] = [
    (1,  "Rotate Image",                "Medium", 48,   "https://leetcode.com/problems/rotate-image/"),
    (2,  "Spiral Matrix",               "Medium", 54,   "https://leetcode.com/problems/spiral-matrix/"),
    (3,  "Set Matrix Zeroes",           "Medium", 73,   "https://leetcode.com/problems/set-matrix-zeroes/"),
    (4,  "Happy Number",                "Easy",   202,  "https://leetcode.com/problems/happy-number/"),
    (5,  "Plus One",                    "Easy",   66,   "https://leetcode.com/problems/plus-one/"),
    (6,  "Pow(x, n)",                   "Medium", 50,   "https://leetcode.com/problems/powx-n/"),
    (7,  "Multiply Strings",            "Medium", 43,   "https://leetcode.com/problems/multiply-strings/"),
    (8,  "Detect Squares",              "Medium", 2013, "https://leetcode.com/problems/detect-squares/"),
    (9,  "Integer to Roman",            "Medium", 12,   "https://leetcode.com/problems/integer-to-roman/"),
    (10, "Roman to Integer",            "Easy",   13,   "https://leetcode.com/problems/roman-to-integer/"),
    (11, "Count Primes",                "Medium", 204,  "https://leetcode.com/problems/count-primes/"),
    (12, "Excel Sheet Column Number",   "Easy",   171,  "https://leetcode.com/problems/excel-sheet-column-number/"),
]

topics["Bit Manipulation"] = [
    (1,  "Single Number",                        "Easy",   136, "https://leetcode.com/problems/single-number/"),
    (2,  "Number of 1 Bits",                     "Easy",   191, "https://leetcode.com/problems/number-of-1-bits/"),
    (3,  "Counting Bits",                        "Easy",   338, "https://leetcode.com/problems/counting-bits/"),
    (4,  "Reverse Bits",                         "Easy",   190, "https://leetcode.com/problems/reverse-bits/"),
    (5,  "Missing Number",                       "Easy",   268, "https://leetcode.com/problems/missing-number/"),
    (6,  "Sum of Two Integers (No +/-)",         "Medium", 371, "https://leetcode.com/problems/sum-of-two-integers/"),
    (7,  "Reverse Integer",                      "Medium", 7,   "https://leetcode.com/problems/reverse-integer/"),
    (8,  "Single Number II",                     "Medium", 137, "https://leetcode.com/problems/single-number-ii/"),
    (9,  "Single Number III",                    "Medium", 260, "https://leetcode.com/problems/single-number-iii/"),
    (10, "Maximum XOR of Two Numbers",           "Medium", 421, "https://leetcode.com/problems/maximum-xor-of-two-numbers-in-an-array/"),
]

topics["String Manipulation"] = [
    (1,  "Longest Common Prefix",                          "Easy",   14,  "https://leetcode.com/problems/longest-common-prefix/"),
    (2,  "Reverse Words in a String",                      "Medium", 151, "https://leetcode.com/problems/reverse-words-in-a-string/"),
    (3,  "String to Integer (atoi)",                       "Medium", 8,   "https://leetcode.com/problems/string-to-integer-atoi/"),
    (4,  "ZigZag Conversion",                              "Medium", 6,   "https://leetcode.com/problems/zigzag-conversion/"),
    (5,  "Count and Say",                                  "Medium", 38,  "https://leetcode.com/problems/count-and-say/"),
    (6,  "Simplify Path",                                  "Medium", 71,  "https://leetcode.com/problems/simplify-path/"),
    (7,  "Basic Calculator II",                            "Medium", 227, "https://leetcode.com/problems/basic-calculator-ii/"),
    (8,  "Basic Calculator",                               "Hard",   224, "https://leetcode.com/problems/basic-calculator/"),
    (9,  "Longest Substring with At Most K Distinct Chars","Medium", 340, "https://leetcode.com/problems/longest-substring-with-at-most-k-distinct-characters/"),
    (10, "Minimum Window Subsequence",                     "Hard",   727, "https://leetcode.com/problems/minimum-window-subsequence/"),
]

topics["Design / OOP"] = [
    (1,  "LRU Cache",                                "Medium", 146, "https://leetcode.com/problems/lru-cache/"),
    (2,  "LFU Cache",                                "Hard",   460, "https://leetcode.com/problems/lfu-cache/"),
    (3,  "Design HashMap",                           "Easy",   706, "https://leetcode.com/problems/design-hashmap/"),
    (4,  "Design HashSet",                           "Easy",   705, "https://leetcode.com/problems/design-hashset/"),
    (5,  "Design Twitter",                           "Medium", 355, "https://leetcode.com/problems/design-twitter/"),
    (6,  "Design Underground System",                "Medium", 1396,"https://leetcode.com/problems/design-underground-system/"),
    (7,  "Snapshot Array",                           "Medium", 1146,"https://leetcode.com/problems/snapshot-array/"),
    (8,  "Design Stack With Increment",              "Medium", 1381,"https://leetcode.com/problems/design-a-stack-with-increment-operation/"),
    (9,  "Implement Stack using Queues",             "Easy",   225, "https://leetcode.com/problems/implement-stack-using-queues/"),
    (10, "Implement Queue using Stacks",             "Easy",   232, "https://leetcode.com/problems/implement-queue-using-stacks/"),
    (11, "Flatten Nested List Iterator",             "Medium", 341, "https://leetcode.com/problems/flatten-nested-list-iterator/"),
    (12, "Serialize and Deserialize BST",            "Medium", 449, "https://leetcode.com/problems/serialize-and-deserialize-bst/"),
    (13, "Range Sum Query — Mutable",                "Medium", 307, "https://leetcode.com/problems/range-sum-query-mutable/"),
    (14, "My Calendar I",                            "Medium", 729, "https://leetcode.com/problems/my-calendar-i/"),
]

# ── Company data ──────────────────────────────────────────────────────────────
companies = {}

companies["Amazon"] = [
    ("Two Sum",                             1,    "https://leetcode.com/problems/two-sum/"),
    ("LRU Cache",                           146,  "https://leetcode.com/problems/lru-cache/"),
    ("Number of Islands",                   200,  "https://leetcode.com/problems/number-of-islands/"),
    ("Merge K Sorted Lists",                23,   "https://leetcode.com/problems/merge-k-sorted-lists/"),
    ("Trapping Rain Water",                 42,   "https://leetcode.com/problems/trapping-rain-water/"),
    ("Word Ladder",                         127,  "https://leetcode.com/problems/word-ladder/"),
    ("Minimum Difficulty of Job Schedule",  1335, "https://leetcode.com/problems/minimum-difficulty-of-a-job-schedule/"),
    ("Reorder Data in Log Files",           937,  "https://leetcode.com/problems/reorder-data-in-log-files/"),
    ("Critical Connections in a Network",   1192, "https://leetcode.com/problems/critical-connections-in-a-network/"),
    ("K Closest Points to Origin",          973,  "https://leetcode.com/problems/k-closest-points-to-origin/"),
    ("Subsets",                             78,   "https://leetcode.com/problems/subsets/"),
    ("Maximum Subarray",                    53,   "https://leetcode.com/problems/maximum-subarray/"),
    ("Copy List with Random Pointer",       138,  "https://leetcode.com/problems/copy-list-with-random-pointer/"),
    ("Design Underground System",           1396, "https://leetcode.com/problems/design-underground-system/"),
    ("Find All Anagrams in a String",       438,  "https://leetcode.com/problems/find-all-anagrams-in-a-string/"),
]

companies["Google"] = [
    ("Median of Two Sorted Arrays",            4,   "https://leetcode.com/problems/median-of-two-sorted-arrays/"),
    ("Word Break II",                          140, "https://leetcode.com/problems/word-break-ii/"),
    ("Minimum Window Substring",               76,  "https://leetcode.com/problems/minimum-window-substring/"),
    ("Regular Expression Matching",            10,  "https://leetcode.com/problems/regular-expression-matching/"),
    ("Serialize and Deserialize Binary Tree",  297, "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/"),
    ("Maximum Subarray",                       53,  "https://leetcode.com/problems/maximum-subarray/"),
    ("N-Queens",                               51,  "https://leetcode.com/problems/n-queens/"),
    ("Wildcard Matching",                      44,  "https://leetcode.com/problems/wildcard-matching/"),
    ("Longest Increasing Path in Matrix",      329, "https://leetcode.com/problems/longest-increasing-path-in-a-matrix/"),
    ("Swim in Rising Water",                   778, "https://leetcode.com/problems/swim-in-rising-water/"),
    ("Burst Balloons",                         312, "https://leetcode.com/problems/burst-balloons/"),
    ("Text Justification",                     68,  "https://leetcode.com/problems/text-justification/"),
    ("Minimum Number of Refueling Stops",      871, "https://leetcode.com/problems/minimum-number-of-refueling-stops/"),
    ("Stickers to Spell Word",                 691, "https://leetcode.com/problems/stickers-to-spell-word/"),
    ("Largest Rectangle in Histogram",         84,  "https://leetcode.com/problems/largest-rectangle-in-histogram/"),
]

companies["Meta"] = [
    ("Merge Intervals",                        56,  "https://leetcode.com/problems/merge-intervals/"),
    ("Valid Parentheses",                      20,  "https://leetcode.com/problems/valid-parentheses/"),
    ("Binary Tree Right Side View",            199, "https://leetcode.com/problems/binary-tree-right-side-view/"),
    ("Minimum Remove to Make Valid Parens",    1249,"https://leetcode.com/problems/minimum-remove-to-make-valid-parentheses/"),
    ("Move Zeroes",                            283, "https://leetcode.com/problems/move-zeroes/"),
    ("Product of Array Except Self",           238, "https://leetcode.com/problems/product-of-array-except-self/"),
    ("Subarray Sum Equals K",                  560, "https://leetcode.com/problems/subarray-sum-equals-k/"),
    ("Add Binary",                             67,  "https://leetcode.com/problems/add-binary/"),
    ("Lowest Common Ancestor Binary Tree",     236, "https://leetcode.com/problems/lowest-common-ancestor-of-a-binary-tree/"),
    ("Accounts Merge",                         721, "https://leetcode.com/problems/accounts-merge/"),
    ("Number of Islands",                      200, "https://leetcode.com/problems/number-of-islands/"),
    ("Merge K Sorted Lists",                   23,  "https://leetcode.com/problems/merge-k-sorted-lists/"),
    ("Remove Invalid Parentheses",             301, "https://leetcode.com/problems/remove-invalid-parentheses/"),
    ("Flatten Binary Tree to Linked List",     114, "https://leetcode.com/problems/flatten-binary-tree-to-linked-list/"),
    ("3Sum",                                   15,  "https://leetcode.com/problems/3sum/"),
]

companies["Microsoft"] = [
    ("Reverse Linked List",                   206, "https://leetcode.com/problems/reverse-linked-list/"),
    ("LRU Cache",                             146, "https://leetcode.com/problems/lru-cache/"),
    ("Clone Graph",                           133, "https://leetcode.com/problems/clone-graph/"),
    ("Copy List with Random Pointer",         138, "https://leetcode.com/problems/copy-list-with-random-pointer/"),
    ("Word Search",                           79,  "https://leetcode.com/problems/word-search/"),
    ("Course Schedule",                       207, "https://leetcode.com/problems/course-schedule/"),
    ("Maximum Depth of Binary Tree",          104, "https://leetcode.com/problems/maximum-depth-of-binary-tree/"),
    ("Merge Intervals",                       56,  "https://leetcode.com/problems/merge-intervals/"),
    ("Two Sum",                               1,   "https://leetcode.com/problems/two-sum/"),
    ("Number of Islands",                     200, "https://leetcode.com/problems/number-of-islands/"),
    ("Longest Substring Without Repeating",   3,   "https://leetcode.com/problems/longest-substring-without-repeating-characters/"),
    ("Edit Distance",                         72,  "https://leetcode.com/problems/edit-distance/"),
    ("Spiral Matrix",                         54,  "https://leetcode.com/problems/spiral-matrix/"),
    ("Serialize and Deserialize Binary Tree", 297, "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/"),
]

companies["Nvidia"] = [
    ("Maximum Product Subarray",         152, "https://leetcode.com/problems/maximum-product-subarray/"),
    ("Trapping Rain Water",              42,  "https://leetcode.com/problems/trapping-rain-water/"),
    ("Find Median from Data Stream",     295, "https://leetcode.com/problems/find-median-from-data-stream/"),
    ("Sliding Window Maximum",           239, "https://leetcode.com/problems/sliding-window-maximum/"),
    ("Pow(x, n)",                        50,  "https://leetcode.com/problems/powx-n/"),
    ("Count of Smaller Numbers After Self", 315, "https://leetcode.com/problems/count-of-smaller-numbers-after-self/"),
    ("Longest Consecutive Sequence",     128, "https://leetcode.com/problems/longest-consecutive-sequence/"),
    ("Sum of Two Integers",              371, "https://leetcode.com/problems/sum-of-two-integers/"),
    ("Maximum XOR of Two Numbers",       421, "https://leetcode.com/problems/maximum-xor-of-two-numbers-in-an-array/"),
    ("Kth Largest Element in an Array",  215, "https://leetcode.com/problems/kth-largest-element-in-an-array/"),
]

companies["Adobe"] = [
    ("Rotate Image",                     48,  "https://leetcode.com/problems/rotate-image/"),
    ("Word Search",                      79,  "https://leetcode.com/problems/word-search/"),
    ("Merge Intervals",                  56,  "https://leetcode.com/problems/merge-intervals/"),
    ("Binary Tree Level Order Traversal",102, "https://leetcode.com/problems/binary-tree-level-order-traversal/"),
    ("Trapping Rain Water",              42,  "https://leetcode.com/problems/trapping-rain-water/"),
    ("Implement Trie",                   208, "https://leetcode.com/problems/implement-trie-prefix-tree/"),
    ("Valid Parentheses",                20,  "https://leetcode.com/problems/valid-parentheses/"),
    ("Kth Largest Element in an Array",  215, "https://leetcode.com/problems/kth-largest-element-in-an-array/"),
    ("Maximum Product Subarray",         152, "https://leetcode.com/problems/maximum-product-subarray/"),
    ("Longest Common Subsequence",       1143,"https://leetcode.com/problems/longest-common-subsequence/"),
]

companies["Bloomberg"] = [
    ("Two Sum",                          1,   "https://leetcode.com/problems/two-sum/"),
    ("Valid Parentheses",                20,  "https://leetcode.com/problems/valid-parentheses/"),
    ("Merge Intervals",                  56,  "https://leetcode.com/problems/merge-intervals/"),
    ("LRU Cache",                        146, "https://leetcode.com/problems/lru-cache/"),
    ("Design Underground System",        1396,"https://leetcode.com/problems/design-underground-system/"),
    ("Group Anagrams",                   49,  "https://leetcode.com/problems/group-anagrams/"),
    ("Reverse Linked List",              206, "https://leetcode.com/problems/reverse-linked-list/"),
    ("Longest Consecutive Sequence",     128, "https://leetcode.com/problems/longest-consecutive-sequence/"),
    ("Kth Largest Element in an Array",  215, "https://leetcode.com/problems/kth-largest-element-in-an-array/"),
    ("Evaluate Reverse Polish Notation", 150, "https://leetcode.com/problems/evaluate-reverse-polish-notation/"),
    ("Decode Ways",                      91,  "https://leetcode.com/problems/decode-ways/"),
    ("Basic Calculator",                 224, "https://leetcode.com/problems/basic-calculator/"),
]

companies["Goldman Sachs"] = [
    ("Count of Smaller Numbers After Self", 315, "https://leetcode.com/problems/count-of-smaller-numbers-after-self/"),
    ("Find First and Last Position",        34,  "https://leetcode.com/problems/find-first-and-last-position-of-element-in-sorted-array/"),
    ("Minimum Path Sum",                    64,  "https://leetcode.com/problems/minimum-path-sum/"),
    ("Trapping Rain Water",                 42,  "https://leetcode.com/problems/trapping-rain-water/"),
    ("Largest Rectangle in Histogram",      84,  "https://leetcode.com/problems/largest-rectangle-in-histogram/"),
    ("Best Time Buy Sell Stock III",        123, "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-iii/"),
    ("Maximum Subarray",                    53,  "https://leetcode.com/problems/maximum-subarray/"),
    ("Merge K Sorted Lists",               23,  "https://leetcode.com/problems/merge-k-sorted-lists/"),
    ("Partition Equal Subset Sum",          416, "https://leetcode.com/problems/partition-equal-subset-sum/"),
]

companies["JP Morgan"] = [
    ("Maximum Subarray",                53,  "https://leetcode.com/problems/maximum-subarray/"),
    ("Best Time to Buy and Sell Stock", 121, "https://leetcode.com/problems/best-time-to-buy-and-sell-stock/"),
    ("Best Time Buy Sell Stock III",    123, "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-iii/"),
    ("Minimum Path Sum",                64,  "https://leetcode.com/problems/minimum-path-sum/"),
    ("Coin Change",                     322, "https://leetcode.com/problems/coin-change/"),
    ("LRU Cache",                       146, "https://leetcode.com/problems/lru-cache/"),
    ("Word Break",                      139, "https://leetcode.com/problems/word-break/"),
    ("Network Delay Time",              743, "https://leetcode.com/problems/network-delay-time/"),
    ("Unique Paths",                    62,  "https://leetcode.com/problems/unique-paths/"),
    ("Edit Distance",                   72,  "https://leetcode.com/problems/edit-distance/"),
]

companies["Morgan Stanley"] = [
    ("Maximum Subarray",                53,  "https://leetcode.com/problems/maximum-subarray/"),
    ("Best Time to Buy and Sell Stock", 121, "https://leetcode.com/problems/best-time-to-buy-and-sell-stock/"),
    ("LRU Cache",                       146, "https://leetcode.com/problems/lru-cache/"),
    ("Coin Change",                     322, "https://leetcode.com/problems/coin-change/"),
    ("Minimum Path Sum",                64,  "https://leetcode.com/problems/minimum-path-sum/"),
    ("Trapping Rain Water",             42,  "https://leetcode.com/problems/trapping-rain-water/"),
    ("Longest Increasing Subsequence",  300, "https://leetcode.com/problems/longest-increasing-subsequence/"),
    ("Edit Distance",                   72,  "https://leetcode.com/problems/edit-distance/"),
]

companies["DE Shaw"] = [
    ("Median of Two Sorted Arrays",         4,   "https://leetcode.com/problems/median-of-two-sorted-arrays/"),
    ("Regular Expression Matching",         10,  "https://leetcode.com/problems/regular-expression-matching/"),
    ("Count of Smaller Numbers After Self", 315, "https://leetcode.com/problems/count-of-smaller-numbers-after-self/"),
    ("Maximum Rectangle",                   85,  "https://leetcode.com/problems/maximal-rectangle/"),
    ("Burst Balloons",                      312, "https://leetcode.com/problems/burst-balloons/"),
    ("Reverse Pairs",                       493, "https://leetcode.com/problems/reverse-pairs/"),
    ("Number of Digit One",                 233, "https://leetcode.com/problems/number-of-digit-one/"),
    ("Super Egg Drop",                      887, "https://leetcode.com/problems/super-egg-drop/"),
    ("Range Sum Query — Mutable",           307, "https://leetcode.com/problems/range-sum-query-mutable/"),
    ("Longest Increasing Subsequence",      300, "https://leetcode.com/problems/longest-increasing-subsequence/"),
]

companies["Citadel"] = [
    ("Reverse Pairs",                         493,  "https://leetcode.com/problems/reverse-pairs/"),
    ("Find Median from Data Stream",          295,  "https://leetcode.com/problems/find-median-from-data-stream/"),
    ("Maximum Profit in Job Scheduling",      1235, "https://leetcode.com/problems/maximum-profit-in-job-scheduling/"),
    ("Sliding Window Maximum",                239,  "https://leetcode.com/problems/sliding-window-maximum/"),
    ("Count of Range Sum",                    327,  "https://leetcode.com/problems/count-of-range-sum/"),
    ("Number of Visible People in a Queue",   1944, "https://leetcode.com/problems/number-of-visible-people-in-a-queue/"),
    ("Kth Largest Element in an Array",       215,  "https://leetcode.com/problems/kth-largest-element-in-an-array/"),
    ("Minimum Number of Refueling Stops",     871,  "https://leetcode.com/problems/minimum-number-of-refueling-stops/"),
]

companies["Cisco"] = [
    ("Number of Islands",                  200,  "https://leetcode.com/problems/number-of-islands/"),
    ("Course Schedule",                    207,  "https://leetcode.com/problems/course-schedule/"),
    ("Network Delay Time",                 743,  "https://leetcode.com/problems/network-delay-time/"),
    ("Cheapest Flights Within K Stops",    787,  "https://leetcode.com/problems/cheapest-flights-within-k-stops/"),
    ("Clone Graph",                        133,  "https://leetcode.com/problems/clone-graph/"),
    ("Word Ladder",                        127,  "https://leetcode.com/problems/word-ladder/"),
    ("Find City With Smallest Neighbors",  1334, "https://leetcode.com/problems/find-the-city-with-the-smallest-number-of-neighbors-at-a-threshold-distance/"),
    ("Critical Connections in a Network",  1192, "https://leetcode.com/problems/critical-connections-in-a-network/"),
    ("Longest Substring Without Repeating",3,    "https://leetcode.com/problems/longest-substring-without-repeating-characters/"),
]

companies["Freshworks"] = [
    ("Two Sum",                          1,   "https://leetcode.com/problems/two-sum/"),
    ("Group Anagrams",                   49,  "https://leetcode.com/problems/group-anagrams/"),
    ("Valid Parentheses",                20,  "https://leetcode.com/problems/valid-parentheses/"),
    ("Merge Intervals",                  56,  "https://leetcode.com/problems/merge-intervals/"),
    ("LRU Cache",                        146, "https://leetcode.com/problems/lru-cache/"),
    ("Binary Tree Level Order Traversal",102, "https://leetcode.com/problems/binary-tree-level-order-traversal/"),
    ("Find All Anagrams in a String",    438, "https://leetcode.com/problems/find-all-anagrams-in-a-string/"),
    ("Maximum Subarray",                 53,  "https://leetcode.com/problems/maximum-subarray/"),
]

companies["PayPal"] = [
    ("Two Sum",                        1,   "https://leetcode.com/problems/two-sum/"),
    ("Valid Parentheses",              20,  "https://leetcode.com/problems/valid-parentheses/"),
    ("Merge Intervals",                56,  "https://leetcode.com/problems/merge-intervals/"),
    ("Coin Change",                    322, "https://leetcode.com/problems/coin-change/"),
    ("Word Break",                     139, "https://leetcode.com/problems/word-break/"),
    ("Design HashMap",                 706, "https://leetcode.com/problems/design-hashmap/"),
    ("Longest Palindromic Substring",  5,   "https://leetcode.com/problems/longest-palindromic-substring/"),
    ("Top K Frequent Elements",        347, "https://leetcode.com/problems/top-k-frequent-elements/"),
]

companies["LinkedIn"] = [
    ("Accounts Merge",                    721,  "https://leetcode.com/problems/accounts-merge/"),
    ("Maximum Profit in Job Scheduling",  1235, "https://leetcode.com/problems/maximum-profit-in-job-scheduling/"),
    ("Find Median from Data Stream",      295,  "https://leetcode.com/problems/find-median-from-data-stream/"),
    ("Nested List Weight Sum",            339,  "https://leetcode.com/problems/nested-list-weight-sum/"),
    ("Nested List Weight Sum II",         364,  "https://leetcode.com/problems/nested-list-weight-sum-ii/"),
    ("Serialize Deserialize Binary Tree", 297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/"),
    ("Max Points on a Line",              149,  "https://leetcode.com/problems/max-points-on-a-line/"),
    ("Paint House II",                    265,  "https://leetcode.com/problems/paint-house-ii/"),
]

companies["Databricks"] = [
    ("LRU Cache",                        146,  "https://leetcode.com/problems/lru-cache/"),
    ("LFU Cache",                        460,  "https://leetcode.com/problems/lfu-cache/"),
    ("Snapshot Array",                   1146, "https://leetcode.com/problems/snapshot-array/"),
    ("Maximum Profit in Job Scheduling", 1235, "https://leetcode.com/problems/maximum-profit-in-job-scheduling/"),
    ("Minimum Interval per Query",       1851, "https://leetcode.com/problems/minimum-interval-to-include-each-query/"),
    ("Serialize Deserialize Binary Tree",297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/"),
    ("Design Underground System",        1396, "https://leetcode.com/problems/design-underground-system/"),
    ("Decode Ways",                      91,   "https://leetcode.com/problems/decode-ways/"),
]

companies["Netflix"] = [
    ("LRU Cache",                      146, "https://leetcode.com/problems/lru-cache/"),
    ("Task Scheduler",                 621, "https://leetcode.com/problems/task-scheduler/"),
    ("Word Break II",                  140, "https://leetcode.com/problems/word-break-ii/"),
    ("Find Median from Data Stream",   295, "https://leetcode.com/problems/find-median-from-data-stream/"),
    ("Design Twitter",                 355, "https://leetcode.com/problems/design-twitter/"),
    ("Reconstruct Itinerary",          332, "https://leetcode.com/problems/reconstruct-itinerary/"),
    ("Merge Intervals",                56,  "https://leetcode.com/problems/merge-intervals/"),
]

companies["Salesforce"] = [
    ("Two Sum",                    1,    "https://leetcode.com/problems/two-sum/"),
    ("Valid Parentheses",          20,   "https://leetcode.com/problems/valid-parentheses/"),
    ("LRU Cache",                  146,  "https://leetcode.com/problems/lru-cache/"),
    ("Merge Intervals",            56,   "https://leetcode.com/problems/merge-intervals/"),
    ("Course Schedule",            207,  "https://leetcode.com/problems/course-schedule/"),
    ("Design Underground System",  1396, "https://leetcode.com/problems/design-underground-system/"),
    ("Number of Islands",          200,  "https://leetcode.com/problems/number-of-islands/"),
    ("Top K Frequent Elements",    347,  "https://leetcode.com/problems/top-k-frequent-elements/"),
]

companies["Oracle"] = [
    ("Serialize Deserialize Binary Tree", 297, "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/"),
    ("Maximum Subarray",                  53,  "https://leetcode.com/problems/maximum-subarray/"),
    ("Word Break",                        139, "https://leetcode.com/problems/word-break/"),
    ("Clone Graph",                       133, "https://leetcode.com/problems/clone-graph/"),
    ("Edit Distance",                     72,  "https://leetcode.com/problems/edit-distance/"),
    ("Course Schedule II",                210, "https://leetcode.com/problems/course-schedule-ii/"),
    ("Coin Change",                       322, "https://leetcode.com/problems/coin-change/"),
    ("Insert Interval",                   57,  "https://leetcode.com/problems/insert-interval/"),
]

companies["SAP"] = [
    ("Two Sum",                          1,    "https://leetcode.com/problems/two-sum/"),
    ("Longest Common Subsequence",       1143, "https://leetcode.com/problems/longest-common-subsequence/"),
    ("Binary Tree Level Order Traversal",102,  "https://leetcode.com/problems/binary-tree-level-order-traversal/"),
    ("Maximum Subarray",                 53,   "https://leetcode.com/problems/maximum-subarray/"),
    ("Merge Intervals",                  56,   "https://leetcode.com/problems/merge-intervals/"),
    ("Validate Binary Search Tree",      98,   "https://leetcode.com/problems/validate-binary-search-tree/"),
    ("Top K Frequent Elements",          347,  "https://leetcode.com/problems/top-k-frequent-elements/"),
]

companies["Atlassian"] = [
    ("LRU Cache",                    146,  "https://leetcode.com/problems/lru-cache/"),
    ("Design Underground System",    1396, "https://leetcode.com/problems/design-underground-system/"),
    ("Merge Intervals",              56,   "https://leetcode.com/problems/merge-intervals/"),
    ("Valid Parentheses",            20,   "https://leetcode.com/problems/valid-parentheses/"),
    ("Course Schedule",              207,  "https://leetcode.com/problems/course-schedule/"),
    ("Word Search",                  79,   "https://leetcode.com/problems/word-search/"),
    ("K Closest Points to Origin",   973,  "https://leetcode.com/problems/k-closest-points-to-origin/"),
]

companies["Barclays"] = [
    ("Best Time to Buy and Sell Stock",    121, "https://leetcode.com/problems/best-time-to-buy-and-sell-stock/"),
    ("Best Time Buy Sell Stock III",       123, "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-iii/"),
    ("Best Time Buy Sell Stock IV",        188, "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-iv/"),
    ("Maximum Subarray",                   53,  "https://leetcode.com/problems/maximum-subarray/"),
    ("Coin Change",                        322, "https://leetcode.com/problems/coin-change/"),
    ("LRU Cache",                          146, "https://leetcode.com/problems/lru-cache/"),
    ("Network Delay Time",                 743, "https://leetcode.com/problems/network-delay-time/"),
    ("Minimum Path Sum",                   64,  "https://leetcode.com/problems/minimum-path-sum/"),
    ("Partition Equal Subset Sum",         416, "https://leetcode.com/problems/partition-equal-subset-sum/"),
]

companies["BlackRock"] = [
    ("Best Time to Buy and Sell Stock",    121, "https://leetcode.com/problems/best-time-to-buy-and-sell-stock/"),
    ("Best Time Buy Sell Stock III",       123, "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-iii/"),
    ("Maximum Subarray",                   53,  "https://leetcode.com/problems/maximum-subarray/"),
    ("LRU Cache",                          146, "https://leetcode.com/problems/lru-cache/"),
    ("Coin Change",                        322, "https://leetcode.com/problems/coin-change/"),
    ("Partition Equal Subset Sum",         416, "https://leetcode.com/problems/partition-equal-subset-sum/"),
    ("Network Delay Time",                 743, "https://leetcode.com/problems/network-delay-time/"),
    ("Minimum Path Sum",                   64,  "https://leetcode.com/problems/minimum-path-sum/"),
]

companies["American Express"] = [
    ("Best Time to Buy and Sell Stock",    121, "https://leetcode.com/problems/best-time-to-buy-and-sell-stock/"),
    ("Maximum Subarray",                   53,  "https://leetcode.com/problems/maximum-subarray/"),
    ("Coin Change",                        322, "https://leetcode.com/problems/coin-change/"),
    ("LRU Cache",                          146, "https://leetcode.com/problems/lru-cache/"),
    ("Minimum Path Sum",                   64,  "https://leetcode.com/problems/minimum-path-sum/"),
    ("Partition Equal Subset Sum",         416, "https://leetcode.com/problems/partition-equal-subset-sum/"),
    ("Two Sum",                            1,   "https://leetcode.com/problems/two-sum/"),
    ("Merge Intervals",                    56,  "https://leetcode.com/problems/merge-intervals/"),
]

companies["Spotify"] = [
    ("Design Twitter",                355, "https://leetcode.com/problems/design-twitter/"),
    ("Top K Frequent Elements",       347, "https://leetcode.com/problems/top-k-frequent-elements/"),
    ("Task Scheduler",                621, "https://leetcode.com/problems/task-scheduler/"),
    ("Accounts Merge",                721, "https://leetcode.com/problems/accounts-merge/"),
    ("Word Ladder",                   127, "https://leetcode.com/problems/word-ladder/"),
    ("LRU Cache",                     146, "https://leetcode.com/problems/lru-cache/"),
    ("K Closest Points to Origin",    973, "https://leetcode.com/problems/k-closest-points-to-origin/"),
]

companies["Dell"] = [
    ("LRU Cache",                        146,  "https://leetcode.com/problems/lru-cache/"),
    ("LFU Cache",                        460,  "https://leetcode.com/problems/lfu-cache/"),
    ("Number of Islands",                200,  "https://leetcode.com/problems/number-of-islands/"),
    ("Clone Graph",                      133,  "https://leetcode.com/problems/clone-graph/"),
    ("Critical Connections in Network",  1192, "https://leetcode.com/problems/critical-connections-in-a-network/"),
    ("Serialize Deserialize Binary Tree",297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/"),
    ("Minimum Difficulty of Job Schedule",1335,"https://leetcode.com/problems/minimum-difficulty-of-a-job-schedule/"),
]

companies["Intel"] = [
    ("LRU Cache",                        146,  "https://leetcode.com/problems/lru-cache/"),
    ("LFU Cache",                        460,  "https://leetcode.com/problems/lfu-cache/"),
    ("Number of Islands",                200,  "https://leetcode.com/problems/number-of-islands/"),
    ("Clone Graph",                      133,  "https://leetcode.com/problems/clone-graph/"),
    ("Critical Connections in Network",  1192, "https://leetcode.com/problems/critical-connections-in-a-network/"),
    ("Maximum Product Subarray",         152,  "https://leetcode.com/problems/maximum-product-subarray/"),
    ("Sliding Window Maximum",           239,  "https://leetcode.com/problems/sliding-window-maximum/"),
]

companies["VMware"] = [
    ("LRU Cache",                        146,  "https://leetcode.com/problems/lru-cache/"),
    ("LFU Cache",                        460,  "https://leetcode.com/problems/lfu-cache/"),
    ("Clone Graph",                      133,  "https://leetcode.com/problems/clone-graph/"),
    ("Critical Connections in Network",  1192, "https://leetcode.com/problems/critical-connections-in-a-network/"),
    ("Serialize Deserialize Binary Tree",297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/"),
    ("Design Underground System",        1396, "https://leetcode.com/problems/design-underground-system/"),
]

companies["Nutanix"] = [
    ("LRU Cache",                        146,  "https://leetcode.com/problems/lru-cache/"),
    ("LFU Cache",                        460,  "https://leetcode.com/problems/lfu-cache/"),
    ("Critical Connections in Network",  1192, "https://leetcode.com/problems/critical-connections-in-a-network/"),
    ("Number of Islands",                200,  "https://leetcode.com/problems/number-of-islands/"),
    ("Serialize Deserialize Binary Tree",297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/"),
    ("Minimum Difficulty of Job Schedule",1335,"https://leetcode.com/problems/minimum-difficulty-of-a-job-schedule/"),
]

companies["Pure Storage"] = [
    ("LRU Cache",                        146,  "https://leetcode.com/problems/lru-cache/"),
    ("LFU Cache",                        460,  "https://leetcode.com/problems/lfu-cache/"),
    ("Critical Connections in Network",  1192, "https://leetcode.com/problems/critical-connections-in-a-network/"),
    ("Sliding Window Maximum",           239,  "https://leetcode.com/problems/sliding-window-maximum/"),
    ("Find Median from Data Stream",     295,  "https://leetcode.com/problems/find-median-from-data-stream/"),
    ("Range Sum Query — Mutable",        307,  "https://leetcode.com/problems/range-sum-query-mutable/"),
]

companies["MathWorks"] = [
    ("Spiral Matrix",                54,  "https://leetcode.com/problems/spiral-matrix/"),
    ("Rotate Image",                 48,  "https://leetcode.com/problems/rotate-image/"),
    ("Set Matrix Zeroes",            73,  "https://leetcode.com/problems/set-matrix-zeroes/"),
    ("Pow(x, n)",                    50,  "https://leetcode.com/problems/powx-n/"),
    ("Multiply Strings",             43,  "https://leetcode.com/problems/multiply-strings/"),
    ("Maximal Square",               221, "https://leetcode.com/problems/maximal-square/"),
    ("Number of 1 Bits",             191, "https://leetcode.com/problems/number-of-1-bits/"),
    ("Count Primes",                 204, "https://leetcode.com/problems/count-primes/"),
    ("Search a 2D Matrix",           74,  "https://leetcode.com/problems/search-a-2d-matrix/"),
]

companies["McKinsey"] = [
    ("Coin Change",                  322, "https://leetcode.com/problems/coin-change/"),
    ("Word Break",                   139, "https://leetcode.com/problems/word-break/"),
    ("Course Schedule",              207, "https://leetcode.com/problems/course-schedule/"),
    ("Minimum Path Sum",             64,  "https://leetcode.com/problems/minimum-path-sum/"),
    ("LRU Cache",                    146, "https://leetcode.com/problems/lru-cache/"),
    ("Partition Equal Subset Sum",   416, "https://leetcode.com/problems/partition-equal-subset-sum/"),
]

# ═══════════════════════════════════════════════════════════════════════════════
#  BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════

# Topic sheets
for topic_name, rows in topics.items():
    add_topic_sheet(wb, topic_name, rows)

# Company sheets
for company_name, rows in companies.items():
    add_company_sheet(wb, company_name, rows)

# ── Study Plan sheet ──────────────────────────────────────────────────────────
ws_plan = wb.create_sheet(title="Study Plan")

plan_data = [
    ("Phase 1 — Warm-up",          "Weeks 1–2",   "Arrays, Hashing, Two Pointers, Sliding Window, Stack",                   "~5 problems/day | Solve all Easy problems in each topic"),
    ("Phase 2 — Core Strengthening","Weeks 3–6",  "Trees, Graphs, Binary Search, Linked Lists, Backtracking, Heaps",        "~4–5 problems/day | Complete full topic blocks"),
    ("Phase 3 — Hard & Companies",  "Weeks 7–10", "DP (1D & 2D), Advanced Graphs, Tries, Company-specific lists",           "~3–4 problems/day | Review solutions after each attempt"),
    ("Phase 4 — Mock Interviews",   "Weeks 11–12","Full timed practice: LeetCode Contest, Pramp, interviewing.io",          "Revisit all Hard problems; do 2+ mocks per week"),
]

resources = [
    ("NeetCode Roadmap",        "https://neetcode.io/roadmap"),
    ("Blind 75",                "https://leetcode.com/discuss/general-discussion/460599/blind-75-leetcode-questions"),
    ("Grind 75",                "https://www.techinterviewhandbook.org/grind75"),
    ("Top Interview 150",       "https://leetcode.com/studyplan/top-interview-150/"),
    ("System Design Primer",    "https://github.com/donnemartin/system-design-primer"),
    ("Pramp (Mock Interviews)", "https://www.pramp.com"),
    ("interviewing.io",         "https://interviewing.io"),
]

ws_plan.merge_cells("A1:D1")
c = ws_plan["A1"]
c.value = "12-Week DSA Study Plan"
c.fill  = HEADER_FILLS["plan"]
c.font  = Font(color="FFFFFF", bold=True, size=14)
c.alignment = CENTER
ws_plan.row_dimensions[1].height = 32

phase_hdrs = ["Phase", "Timeline", "Topics", "Daily Target"]
phase_widths = [28, 14, 60, 55]
for col, (h, w) in enumerate(zip(phase_hdrs, phase_widths), 1):
    cell = ws_plan.cell(row=2, column=col, value=h)
    cell.fill      = HEADER_FILLS["col_hdr"]
    cell.font      = WHITE
    cell.alignment = CENTER
    cell.border    = BOX
    ws_plan.column_dimensions[get_column_letter(col)].width = w
ws_plan.row_dimensions[2].height = 20

PHASE_COLORS = ["E3F2FD", "E8F5E9", "FFF8E1", "FCE4EC"]
for r_idx, (phase, timeline, topics_str, target) in enumerate(plan_data, start=3):
    fill = PatternFill("solid", fgColor=PHASE_COLORS[r_idx - 3])
    for col, val in enumerate([phase, timeline, topics_str, target], start=1):
        cell = ws_plan.cell(row=r_idx, column=col, value=val)
        cell.fill      = fill
        cell.font      = BOLD if col == 1 else NORMAL
        cell.alignment = LEFT
        cell.border    = BOX
    ws_plan.row_dimensions[r_idx].height = 30

# Resources section
ws_plan.cell(row=8, column=1, value="").fill = PatternFill("solid", fgColor="FFFFFF")

ws_plan.merge_cells("A9:D9")
c9 = ws_plan["A9"]
c9.value = "Useful Resources"
c9.fill  = HEADER_FILLS["plan"]
c9.font  = Font(color="FFFFFF", bold=True, size=12)
c9.alignment = CENTER
ws_plan.row_dimensions[9].height = 24

ws_plan.cell(row=10, column=1, value="Resource").fill = HEADER_FILLS["col_hdr"]
ws_plan.cell(row=10, column=1).font      = WHITE
ws_plan.cell(row=10, column=1).alignment = CENTER
ws_plan.cell(row=10, column=1).border    = BOX

ws_plan.merge_cells("B10:D10")
ws_plan.cell(row=10, column=2, value="Link").fill = HEADER_FILLS["col_hdr"]
ws_plan.cell(row=10, column=2).font      = WHITE
ws_plan.cell(row=10, column=2).alignment = CENTER
ws_plan.cell(row=10, column=2).border    = BOX
ws_plan.row_dimensions[10].height = 20

for r_idx, (name, url) in enumerate(resources, start=11):
    fill = PatternFill("solid", fgColor="E8EAF6" if r_idx % 2 == 0 else "FFFFFF")
    ws_plan.cell(row=r_idx, column=1, value=name).fill      = fill
    ws_plan.cell(row=r_idx, column=1).font      = BOLD
    ws_plan.cell(row=r_idx, column=1).alignment = LEFT
    ws_plan.cell(row=r_idx, column=1).border    = BOX

    ws_plan.merge_cells(f"B{r_idx}:D{r_idx}")
    lc = ws_plan.cell(row=r_idx, column=2, value=url)
    lc.hyperlink  = url
    lc.font       = LINK
    lc.fill       = fill
    lc.alignment  = LEFT
    lc.border     = BOX
    ws_plan.row_dimensions[r_idx].height = 18

ws_plan.freeze_panes = "A3"

# ── Index / Overview sheet (first sheet) ──────────────────────────────────────
ws_idx = wb.create_sheet(title="Overview", index=0)

ws_idx.merge_cells("A1:C1")
c = ws_idx["A1"]
c.value = "DSA Interview Prep — Overview"
c.fill  = HEADER_FILLS["title"]
c.font  = Font(color="FFFFFF", bold=True, size=14)
c.alignment = CENTER
ws_idx.row_dimensions[1].height = 32

ws_idx.column_dimensions["A"].width = 30
ws_idx.column_dimensions["B"].width = 18
ws_idx.column_dimensions["C"].width = 40

for col, h in enumerate(["Sheet Name", "# Questions", "Description"], start=1):
    cell = ws_idx.cell(row=2, column=col, value=h)
    cell.fill      = HEADER_FILLS["col_hdr"]
    cell.font      = WHITE
    cell.alignment = CENTER
    cell.border    = BOX
ws_idx.row_dimensions[2].height = 20

overview_rows = [
    ("Arrays & Hashing",      len(topics["Arrays & Hashing"]),      "Hash maps, frequency counting, prefix sums"),
    ("Two Pointers",           len(topics["Two Pointers"]),           "Sorted arrays, palindrome, opposite ends"),
    ("Sliding Window",         len(topics["Sliding Window"]),         "Variable/fixed size windows on arrays/strings"),
    ("Stack",                  len(topics["Stack"]),                  "Monotonic stack, parentheses, histograms"),
    ("Binary Search",          len(topics["Binary Search"]),          "Search on sorted space, binary search on answer"),
    ("Linked List",            len(topics["Linked List"]),            "Reversal, cycle detection, merge, pointers"),
    ("Trees",                  len(topics["Trees"]),                  "DFS/BFS on binary trees, BST operations"),
    ("Heap / Priority Queue",  len(topics["Heap / Priority Queue"]),  "Top-K, median, scheduling problems"),
    ("Backtracking",           len(topics["Backtracking"]),           "Subsets, permutations, pruning decision trees"),
    ("Tries",                  len(topics["Tries"]),                  "Prefix tree, word search, XOR trie"),
    ("Graphs",                 len(topics["Graphs"]),                 "BFS/DFS, Union-Find, topological sort"),
    ("Advanced Graphs",        len(topics["Advanced Graphs"]),        "Dijkstra, Bellman-Ford, Prim, Kruskal"),
    ("DP — 1D",                len(topics["DP — 1D"]),                "Linear DP: LIS, coin change, house robber"),
    ("DP — 2D",                len(topics["DP — 2D"]),                "Grid DP, interval DP, string DP"),
    ("Greedy",                 len(topics["Greedy"]),                 "Local optimal choices leading to global optimum"),
    ("Intervals",              len(topics["Intervals"]),              "Merge, insert, sweep line"),
    ("Math & Geometry",        len(topics["Math & Geometry"]),        "Matrix ops, number theory, geometry"),
    ("Bit Manipulation",       len(topics["Bit Manipulation"]),       "XOR tricks, bit masking, counting bits"),
    ("String Manipulation",    len(topics["String Manipulation"]),    "Parsing, pattern matching, calculator"),
    ("Design / OOP",           len(topics["Design / OOP"]),           "LRU/LFU cache, iterators, data structure design"),
]

for cname in companies:
    overview_rows.append((cname, len(companies[cname]), "Company-specific high-frequency questions"))

overview_rows.append(("Study Plan", "-", "12-week preparation roadmap + resources"))

TOPIC_FILL  = PatternFill("solid", fgColor="E3F2FD")
COMP_FILL   = PatternFill("solid", fgColor="F3E5F5")
PLAN_FILL   = PatternFill("solid", fgColor="E8F5E9")

for r_idx, (name, count, desc) in enumerate(overview_rows, start=3):
    if name in topics:
        fill = TOPIC_FILL
    elif name == "Study Plan":
        fill = PLAN_FILL
    else:
        fill = COMP_FILL

    ws_idx.cell(row=r_idx, column=1, value=name).fill      = fill
    ws_idx.cell(row=r_idx, column=1).font      = BOLD
    ws_idx.cell(row=r_idx, column=1).alignment = LEFT
    ws_idx.cell(row=r_idx, column=1).border    = BOX

    ws_idx.cell(row=r_idx, column=2, value=count).fill      = fill
    ws_idx.cell(row=r_idx, column=2).font      = NORMAL
    ws_idx.cell(row=r_idx, column=2).alignment = CENTER
    ws_idx.cell(row=r_idx, column=2).border    = BOX

    ws_idx.cell(row=r_idx, column=3, value=desc).fill      = fill
    ws_idx.cell(row=r_idx, column=3).font      = NORMAL
    ws_idx.cell(row=r_idx, column=3).alignment = LEFT
    ws_idx.cell(row=r_idx, column=3).border    = BOX
    ws_idx.row_dimensions[r_idx].height = 18

ws_idx.freeze_panes = "A3"

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = r"c:\Users\MaitySom\Desktop\Somneel\Somneel_Maity\K_Tech\DSA\DSA_Interview_Prep.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Sheets created: {len(wb.sheetnames)}")
