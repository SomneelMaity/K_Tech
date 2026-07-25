"""
generate_google_dsa_excel.py
────────────────────────────
Generates a Google-focused DSA preparation Excel workbook.
Sheets:
  0 – Overview          : category summary + Google interview tips
  1…N – Category sheets : questions organised by topic, with status tracker
  Last – Study Plan     : Google-specific 8-week plan
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Workbook ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ─── Palette ──────────────────────────────────────────────────────────────────
GOOGLE_BLUE   = "4285F4"
GOOGLE_RED    = "EA4335"
GOOGLE_YELLOW = "FBBC05"
GOOGLE_GREEN  = "34A853"

FILLS = {
    "title"  : PatternFill("solid", fgColor="1A237E"),   # deep indigo
    "col_hdr": PatternFill("solid", fgColor=GOOGLE_BLUE),
    "easy"   : PatternFill("solid", fgColor="C8E6C9"),   # light green
    "medium" : PatternFill("solid", fgColor="FFF9C4"),   # light yellow
    "hard"   : PatternFill("solid", fgColor="FFCDD2"),   # light red
    "section": PatternFill("solid", fgColor="E8EAF6"),   # light indigo
    "google" : PatternFill("solid", fgColor="E3F2FD"),   # light blue
    "plan"   : PatternFill("solid", fgColor="00695C"),   # teal
    "tip"    : PatternFill("solid", fgColor="FFF8E1"),   # warm yellow
}

DIFF_FILL = {
    "Easy"  : FILLS["easy"],
    "Medium": FILLS["medium"],
    "Hard"  : FILLS["hard"],
}

WHITE  = Font(color="FFFFFF", bold=True, size=11)
BOLD   = Font(bold=True, size=10)
NORMAL = Font(size=10)
LINK   = Font(color=GOOGLE_BLUE, underline="single", size=10)
STRIKE = Font(size=10, strike=True, color="9E9E9E")   # for solved items
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN = Side(style="thin", color="BDBDBD")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ─── Helpers ──────────────────────────────────────────────────────────────────
def safe_title(name):
    for ch in r"/\?*[]:'":
        name = name.replace(ch, "-")
    return name[:31]


def add_category_sheet(wb, category, description, rows):
    """
    rows: list of (no, problem, difficulty, lc_number, url, notes)
    Adds: #, Problem, Difficulty, LC#, LeetCode Link, Notes, Status
    """
    ws = wb.create_sheet(title=safe_title(category))

    # ── Title ──
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = f"Google DSA Prep — {category}"
    c.fill      = FILLS["title"]
    c.font      = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # ── Description row ──
    ws.merge_cells("A2:G2")
    d = ws["A2"]
    d.value     = description
    d.fill      = FILLS["tip"]
    d.font      = Font(italic=True, size=10, color="333333")
    d.alignment = LEFT
    ws.row_dimensions[2].height = 18

    # ── Column headers ──
    headers    = ["#", "Problem", "Difficulty", "LC #", "LeetCode Link", "Notes / Pattern", "Status"]
    col_widths = [5,   50,        12,           8,      46,              38,                 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill      = FILLS["col_hdr"]
        cell.font      = WHITE
        cell.alignment = CENTER
        cell.border    = BOX
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 20

    # ── Data rows ──
    for r_idx, (no, problem, difficulty, lc_num, url, notes) in enumerate(rows, start=4):
        fill = DIFF_FILL.get(difficulty, PatternFill())

        ws.cell(row=r_idx, column=1, value=no).alignment        = CENTER
        ws.cell(row=r_idx, column=2, value=problem).alignment   = LEFT
        ws.cell(row=r_idx, column=3, value=difficulty).alignment = CENTER
        ws.cell(row=r_idx, column=4, value=lc_num).alignment    = CENTER

        link_cell = ws.cell(row=r_idx, column=5, value=f"LC {lc_num} — Open")
        link_cell.hyperlink  = url
        link_cell.font       = LINK
        link_cell.alignment  = LEFT

        ws.cell(row=r_idx, column=6, value=notes).alignment = LEFT
        ws.cell(row=r_idx, column=6).font = Font(size=9, color="555555", italic=True)

        status_cell = ws.cell(row=r_idx, column=7, value="⬜ To Do")
        status_cell.alignment = CENTER
        status_cell.font      = Font(size=9, color="555555")

        for col in range(1, 8):
            cell = ws.cell(row=r_idx, column=col)
            cell.fill   = fill
            cell.border = BOX
            if col not in (5, 6, 7):
                cell.font = NORMAL
        ws.row_dimensions[r_idx].height = 18

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:G{3 + len(rows)}"
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
#  GOOGLE DSA QUESTION BANK — Organised by Category
#  Format: (no, problem_name, difficulty, lc_number, url, pattern_notes)
# ═══════════════════════════════════════════════════════════════════════════════

categories = {}

# ── 1. Arrays & Hashing ───────────────────────────────────────────────────────
categories["Arrays & Hashing"] = (
    "Google focus: hash maps, prefix sums, counting, range queries. Expect O(n) or O(n log n) solutions.",
    [
        (1,  "Two Sum",                                 "Easy",   1,    "https://leetcode.com/problems/two-sum/",                                          "Hash map; complement lookup"),
        (2,  "Contains Duplicate",                      "Easy",   217,  "https://leetcode.com/problems/contains-duplicate/",                               "Hash set"),
        (3,  "Group Anagrams",                          "Medium", 49,   "https://leetcode.com/problems/group-anagrams/",                                    "Sorted key or char-count tuple"),
        (4,  "Top K Frequent Elements",                 "Medium", 347,  "https://leetcode.com/problems/top-k-frequent-elements/",                          "Heap or bucket sort"),
        (5,  "Product of Array Except Self",            "Medium", 238,  "https://leetcode.com/problems/product-of-array-except-self/",                     "Prefix & suffix products; no division"),
        (6,  "Longest Consecutive Sequence",            "Medium", 128,  "https://leetcode.com/problems/longest-consecutive-sequence/",                     "Hash set; O(n)"),
        (7,  "Subarray Sum Equals K",                   "Medium", 560,  "https://leetcode.com/problems/subarray-sum-equals-k/",                            "Prefix sum + hash map"),
        (8,  "Sort Colors (Dutch Flag)",                "Medium", 75,   "https://leetcode.com/problems/sort-colors/",                                       "3-pointer in-place"),
        (9,  "Find All Duplicates in an Array",         "Medium", 442,  "https://leetcode.com/problems/find-all-duplicates-in-an-array/",                  "Negate-at-index trick"),
        (10, "Maximum Points You Can Obtain from Cards","Medium", 1423, "https://leetcode.com/problems/maximum-points-you-can-obtain-from-cards/",         "Sliding window (fixed)"),
        (11, "First Missing Positive",                  "Hard",   41,   "https://leetcode.com/problems/first-missing-positive/",                           "Index-as-hash; cyclic sort"),
        (12, "Trapping Rain Water",                     "Hard",   42,   "https://leetcode.com/problems/trapping-rain-water/",                               "Two pointers or monotonic stack"),
        (13, "Largest Rectangle in Histogram",          "Hard",   84,   "https://leetcode.com/problems/largest-rectangle-in-histogram/",                   "Monotonic stack"),
        (14, "Maximum Rectangle",                       "Hard",   85,   "https://leetcode.com/problems/maximal-rectangle/",                                 "Histogram DP per row"),
        (15, "Count of Smaller Numbers After Self",     "Hard",   315,  "https://leetcode.com/problems/count-of-smaller-numbers-after-self/",              "Merge sort / BIT / Segment tree"),
    ]
)

# ── 2. Strings & Sliding Window ───────────────────────────────────────────────
categories["Strings & Sliding Window"] = (
    "Google frequently tests string parsing, pattern matching, and window-based optimisation. Master variable-size windows.",
    [
        (1,  "Longest Substring Without Repeating Chars",      "Medium", 3,   "https://leetcode.com/problems/longest-substring-without-repeating-characters/", "Sliding window + set"),
        (2,  "Minimum Window Substring",                       "Hard",   76,  "https://leetcode.com/problems/minimum-window-substring/",                        "Sliding window + char count"),
        (3,  "Longest Repeating Character Replacement",        "Medium", 424, "https://leetcode.com/problems/longest-repeating-character-replacement/",         "Sliding window; max-freq trick"),
        (4,  "Permutation in String",                          "Medium", 567, "https://leetcode.com/problems/permutation-in-string/",                           "Fixed window + char count"),
        (5,  "Find All Anagrams in a String",                  "Medium", 438, "https://leetcode.com/problems/find-all-anagrams-in-a-string/",                  "Fixed sliding window"),
        (6,  "Sliding Window Maximum",                         "Hard",   239, "https://leetcode.com/problems/sliding-window-maximum/",                          "Monotonic deque"),
        (7,  "Longest Palindromic Substring",                  "Medium", 5,   "https://leetcode.com/problems/longest-palindromic-substring/",                  "Expand around centre / DP"),
        (8,  "Palindromic Substrings",                         "Medium", 647, "https://leetcode.com/problems/palindromic-substrings/",                          "Expand around centre"),
        (9,  "Minimum Size Subarray Sum",                      "Medium", 209, "https://leetcode.com/problems/minimum-size-subarray-sum/",                      "Variable window"),
        (10, "Text Justification",                             "Hard",   68,  "https://leetcode.com/problems/text-justification/",                              "String simulation; greedy spaces — Google classic"),
        (11, "Valid Anagram",                                  "Easy",   242, "https://leetcode.com/problems/valid-anagram/",                                   "Char frequency"),
        (12, "Decode String",                                  "Medium", 394, "https://leetcode.com/problems/decode-string/",                                   "Stack simulation"),
        (13, "Longest Substring with At Most K Distinct Chars","Medium", 340, "https://leetcode.com/problems/longest-substring-with-at-most-k-distinct-characters/", "Sliding window + map"),
        (14, "Minimum Window Subsequence",                     "Hard",   727, "https://leetcode.com/problems/minimum-window-subsequence/",                     "Two-pointer forward+backward scan"),
        (15, "Word Break",                                     "Medium", 139, "https://leetcode.com/problems/word-break/",                                      "BFS or 1D DP"),
        (16, "Word Break II",                                  "Hard",   140, "https://leetcode.com/problems/word-break-ii/",                                   "Backtracking + memoisation — Google very frequent"),
    ]
)

# ── 3. Two Pointers & Binary Search ───────────────────────────────────────────
categories["Two Pointers & Binary Search"] = (
    "Google loves binary search on the answer, not just on arrays. Think 'what am I searching for?' before coding.",
    [
        (1,  "Median of Two Sorted Arrays",            "Hard",   4,    "https://leetcode.com/problems/median-of-two-sorted-arrays/",                     "Binary search on partition — Google classic"),
        (2,  "3Sum",                                   "Medium", 15,   "https://leetcode.com/problems/3sum/",                                             "Sort + two pointers"),
        (3,  "Container With Most Water",              "Medium", 11,   "https://leetcode.com/problems/container-with-most-water/",                        "Two pointers; greedy shrink"),
        (4,  "Search in Rotated Sorted Array",         "Medium", 33,   "https://leetcode.com/problems/search-in-rotated-sorted-array/",                  "Modified binary search"),
        (5,  "Find Minimum in Rotated Sorted Array",   "Medium", 153,  "https://leetcode.com/problems/find-minimum-in-rotated-sorted-array/",            "Binary search"),
        (6,  "Search a 2D Matrix",                     "Medium", 74,   "https://leetcode.com/problems/search-a-2d-matrix/",                               "Treat as 1D sorted array"),
        (7,  "Koko Eating Bananas",                    "Medium", 875,  "https://leetcode.com/problems/koko-eating-bananas/",                              "Binary search on answer"),
        (8,  "Capacity to Ship Packages Within D Days","Medium", 1011, "https://leetcode.com/problems/capacity-to-ship-packages-within-d-days/",         "Binary search on answer"),
        (9,  "Split Array Largest Sum",                "Hard",   410,  "https://leetcode.com/problems/split-array-largest-sum/",                          "Binary search on answer"),
        (10, "Find First and Last Position",           "Medium", 34,   "https://leetcode.com/problems/find-first-and-last-position-of-element-in-sorted-array/", "Two binary searches"),
        (11, "Path With Minimum Effort",               "Medium", 1631, "https://leetcode.com/problems/path-with-minimum-effort/",                         "Binary search + BFS/DFS"),
        (12, "Minimum Number of Refueling Stops",      "Hard",   871,  "https://leetcode.com/problems/minimum-number-of-refueling-stops/",                "Greedy + max-heap — Google frequent"),
    ]
)

# ── 4. Stack & Monotonic Stack ────────────────────────────────────────────────
categories["Stack & Monotonic Stack"] = (
    "Google tests monotonic stacks for 'next greater/smaller' and histogram problems. Know both increasing and decreasing variants.",
    [
        (1,  "Valid Parentheses",                            "Easy",   20,   "https://leetcode.com/problems/valid-parentheses/",                               "Classic stack"),
        (2,  "Min Stack",                                    "Medium", 155,  "https://leetcode.com/problems/min-stack/",                                        "Pair-stack or auxiliary stack"),
        (3,  "Daily Temperatures",                           "Medium", 739,  "https://leetcode.com/problems/daily-temperatures/",                               "Monotonic decreasing stack"),
        (4,  "Largest Rectangle in Histogram",               "Hard",   84,   "https://leetcode.com/problems/largest-rectangle-in-histogram/",                  "Monotonic increasing stack"),
        (5,  "Trapping Rain Water",                          "Hard",   42,   "https://leetcode.com/problems/trapping-rain-water/",                              "Stack or two-pointer"),
        (6,  "Basic Calculator",                             "Hard",   224,  "https://leetcode.com/problems/basic-calculator/",                                 "Stack for sign context"),
        (7,  "Basic Calculator II",                          "Medium", 227,  "https://leetcode.com/problems/basic-calculator-ii/",                              "Stack; handle * /"),
        (8,  "Minimum Remove to Make Valid Parentheses",     "Medium", 1249, "https://leetcode.com/problems/minimum-remove-to-make-valid-parentheses/",        "Stack of indices"),
        (9,  "Asteroid Collision",                           "Medium", 735,  "https://leetcode.com/problems/asteroid-collision/",                               "Simulate with stack"),
        (10, "Remove Duplicate Letters",                     "Medium", 316,  "https://leetcode.com/problems/remove-duplicate-letters/",                        "Greedy + monotonic stack"),
        (11, "Maximum Width Ramp",                           "Medium", 962,  "https://leetcode.com/problems/maximum-width-ramp/",                               "Monotonic stack then sweep right"),
    ]
)

# ── 5. Linked Lists ───────────────────────────────────────────────────────────
categories["Linked Lists"] = (
    "Google covers linked list pointers, cycle detection (Floyd), merge operations, and in-place reversal.",
    [
        (1,  "Reverse Linked List",                   "Easy",   206, "https://leetcode.com/problems/reverse-linked-list/",                             "Iterative & recursive"),
        (2,  "Merge Two Sorted Lists",                "Easy",   21,  "https://leetcode.com/problems/merge-two-sorted-lists/",                          "Dummy head"),
        (3,  "Linked List Cycle",                     "Easy",   141, "https://leetcode.com/problems/linked-list-cycle/",                               "Floyd's slow/fast"),
        (4,  "Linked List Cycle II",                  "Medium", 142, "https://leetcode.com/problems/linked-list-cycle-ii/",                            "Floyd's + math"),
        (5,  "Remove Nth Node From End",              "Medium", 19,  "https://leetcode.com/problems/remove-nth-node-from-end-of-list/",               "Two pointer with gap n"),
        (6,  "Copy List with Random Pointer",         "Medium", 138, "https://leetcode.com/problems/copy-list-with-random-pointer/",                  "Hash map or weave"),
        (7,  "LRU Cache",                             "Medium", 146, "https://leetcode.com/problems/lru-cache/",                                       "Doubly-LL + hash map"),
        (8,  "Merge K Sorted Lists",                  "Hard",   23,  "https://leetcode.com/problems/merge-k-sorted-lists/",                           "Min-heap or divide & conquer"),
        (9,  "Reverse Nodes in K-Group",              "Hard",   25,  "https://leetcode.com/problems/reverse-nodes-in-k-group/",                       "Reverse in chunks"),
        (10, "Reorder List",                          "Medium", 143, "https://leetcode.com/problems/reorder-list/",                                    "Find middle + reverse + merge"),
        (11, "Sort List",                             "Medium", 148, "https://leetcode.com/problems/sort-list/",                                       "Merge sort on linked list"),
        (12, "Palindrome Linked List",                "Easy",   234, "https://leetcode.com/problems/palindrome-linked-list/",                         "Find mid + reverse second half"),
    ]
)

# ── 6. Trees & BST ────────────────────────────────────────────────────────────
categories["Trees & BST"] = (
    "Google asks deep tree problems: serialisation, vertical traversal, and ancestor queries. Know DFS/BFS variants cold.",
    [
        (1,  "Invert Binary Tree",                           "Easy",   226,  "https://leetcode.com/problems/invert-binary-tree/",                              "Recursive swap"),
        (2,  "Maximum Depth of Binary Tree",                 "Easy",   104,  "https://leetcode.com/problems/maximum-depth-of-binary-tree/",                   "DFS"),
        (3,  "Diameter of Binary Tree",                      "Easy",   543,  "https://leetcode.com/problems/diameter-of-binary-tree/",                        "Height DFS; max left+right"),
        (4,  "Balanced Binary Tree",                         "Easy",   110,  "https://leetcode.com/problems/balanced-binary-tree/",                           "Height check DFS"),
        (5,  "LCA of Binary Tree",                          "Medium", 236,  "https://leetcode.com/problems/lowest-common-ancestor-of-a-binary-tree/",        "Post-order DFS"),
        (6,  "LCA of BST",                                  "Medium", 235,  "https://leetcode.com/problems/lowest-common-ancestor-of-a-binary-search-tree/", "Use BST property"),
        (7,  "Binary Tree Level Order Traversal",            "Medium", 102,  "https://leetcode.com/problems/binary-tree-level-order-traversal/",              "BFS with queue"),
        (8,  "Binary Tree Right Side View",                  "Medium", 199,  "https://leetcode.com/problems/binary-tree-right-side-view/",                    "BFS last per level"),
        (9,  "Binary Tree Maximum Path Sum",                 "Hard",   124,  "https://leetcode.com/problems/binary-tree-maximum-path-sum/",                   "Post-order; global max"),
        (10, "Serialize and Deserialize Binary Tree",        "Hard",   297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/",          "BFS or preorder — Google very frequent"),
        (11, "Vertical Order Traversal of Binary Tree",      "Hard",   987,  "https://leetcode.com/problems/vertical-order-traversal-of-a-binary-tree/",     "DFS with (col,row) keys — Google classic"),
        (12, "All Nodes Distance K in Binary Tree",          "Medium", 863,  "https://leetcode.com/problems/all-nodes-distance-k-in-binary-tree/",           "Graph-ify tree; BFS"),
        (13, "Validate Binary Search Tree",                  "Medium", 98,   "https://leetcode.com/problems/validate-binary-search-tree/",                    "Min/max range DFS"),
        (14, "Kth Smallest Element in BST",                  "Medium", 230,  "https://leetcode.com/problems/kth-smallest-element-in-a-bst/",                 "In-order traversal"),
        (15, "Construct BT from Preorder and Inorder",       "Medium", 105,  "https://leetcode.com/problems/construct-binary-tree-from-preorder-and-inorder-traversal/", "Divide root from sequences"),
        (16, "Recover Binary Search Tree",                   "Medium", 99,   "https://leetcode.com/problems/recover-binary-search-tree/",                    "Morris traversal or in-order + swap"),
        (17, "Path Sum II",                                  "Medium", 113,  "https://leetcode.com/problems/path-sum-ii/",                                     "DFS backtrack"),
        (18, "Count Good Nodes in Binary Tree",              "Medium", 1448, "https://leetcode.com/problems/count-good-nodes-in-binary-tree/",                "DFS; track running max"),
        (19, "Flatten Binary Tree to Linked List",           "Medium", 114,  "https://leetcode.com/problems/flatten-binary-tree-to-linked-list/",             "Morris right-thread"),
        (20, "Populating Next Right Pointers",               "Medium", 116,  "https://leetcode.com/problems/populating-next-right-pointers-in-each-node/",   "BFS level-link"),
    ]
)

# ── 7. Graphs (BFS / DFS / Union-Find / Topological Sort) ────────────────────
categories["Graphs"] = (
    "Google Maps, Search indexing, and dependency resolution — graphs appear everywhere at Google. Master all traversal methods.",
    [
        (1,  "Number of Islands",                        "Medium", 200,  "https://leetcode.com/problems/number-of-islands/",                              "BFS/DFS flood-fill"),
        (2,  "Word Ladder",                              "Hard",   127,  "https://leetcode.com/problems/word-ladder/",                                    "BFS with word transforms — Google classic"),
        (3,  "Clone Graph",                              "Medium", 133,  "https://leetcode.com/problems/clone-graph/",                                    "BFS/DFS + hash map"),
        (4,  "Pacific Atlantic Water Flow",               "Medium", 417,  "https://leetcode.com/problems/pacific-atlantic-water-flow/",                   "Reverse DFS from both oceans"),
        (5,  "Course Schedule",                          "Medium", 207,  "https://leetcode.com/problems/course-schedule/",                                "Topo sort / cycle detect"),
        (6,  "Course Schedule II",                       "Medium", 210,  "https://leetcode.com/problems/course-schedule-ii/",                             "Kahn's BFS topo sort"),
        (7,  "Accounts Merge",                           "Medium", 721,  "https://leetcode.com/problems/accounts-merge/",                                 "Union-Find or DFS — Google frequent"),
        (8,  "Rotting Oranges",                          "Medium", 994,  "https://leetcode.com/problems/rotting-oranges/",                                "Multi-source BFS"),
        (9,  "Is Graph Bipartite?",                      "Medium", 785,  "https://leetcode.com/problems/is-graph-bipartite/",                             "2-colour BFS/DFS"),
        (10, "Find Eventual Safe States",                "Medium", 802,  "https://leetcode.com/problems/find-eventual-safe-states/",                     "Reverse graph topo sort"),
        (11, "Critical Connections in a Network",        "Hard",   1192, "https://leetcode.com/problems/critical-connections-in-a-network/",             "Tarjan's bridge algorithm"),
        (12, "Redundant Connection",                     "Medium", 684,  "https://leetcode.com/problems/redundant-connection/",                           "Union-Find; detect cycle"),
        (13, "Minimum Number of Vertices to Reach All",  "Medium", 1557, "https://leetcode.com/problems/minimum-number-of-vertices-to-reach-all-nodes/", "Nodes with 0 in-degree"),
        (14, "Swim in Rising Water",                     "Hard",   778,  "https://leetcode.com/problems/swim-in-rising-water/",                          "Dijkstra or Binary search + BFS"),
        (15, "Reconstruct Itinerary",                    "Hard",   332,  "https://leetcode.com/problems/reconstruct-itinerary/",                         "Hierholzer's Eulerian path"),
        (16, "Network Delay Time",                       "Medium", 743,  "https://leetcode.com/problems/network-delay-time/",                             "Dijkstra SSSP"),
        (17, "Cheapest Flights Within K Stops",          "Medium", 787,  "https://leetcode.com/problems/cheapest-flights-within-k-stops/",               "Bellman-Ford (k+1 rounds) / BFS"),
        (18, "All Paths From Source to Target",          "Medium", 797,  "https://leetcode.com/problems/all-paths-from-source-to-target/",               "DFS backtrack"),
        (19, "Evaluate Division",                        "Medium", 399,  "https://leetcode.com/problems/evaluate-division/",                              "Weighted graph BFS/DFS"),
        (20, "Shortest Path in Binary Matrix",           "Medium", 1091, "https://leetcode.com/problems/shortest-path-in-binary-matrix/",                "BFS; 8-directional"),
    ]
)

# ── 8. Dynamic Programming ────────────────────────────────────────────────────
categories["Dynamic Programming"] = (
    "Google's most tested hard area. Focus on: interval DP, string DP, game theory DP, path DP. Always analyse overlapping sub-problems.",
    [
        (1,  "Climbing Stairs",                                "Easy",   70,   "https://leetcode.com/problems/climbing-stairs/",                                  "1D DP; Fibonacci"),
        (2,  "House Robber",                                   "Medium", 198,  "https://leetcode.com/problems/house-robber/",                                     "1D DP; no adjacent"),
        (3,  "Coin Change",                                    "Medium", 322,  "https://leetcode.com/problems/coin-change/",                                      "Unbounded knapsack"),
        (4,  "Longest Increasing Subsequence",                 "Medium", 300,  "https://leetcode.com/problems/longest-increasing-subsequence/",                  "1D DP or patience sort O(n log n)"),
        (5,  "Longest Common Subsequence",                     "Medium", 1143, "https://leetcode.com/problems/longest-common-subsequence/",                      "2D DP"),
        (6,  "Edit Distance",                                  "Medium", 72,   "https://leetcode.com/problems/edit-distance/",                                    "2D DP; classic string DP"),
        (7,  "Regular Expression Matching",                    "Hard",   10,   "https://leetcode.com/problems/regular-expression-matching/",                     "2D DP with '*' lookahead — Google classic"),
        (8,  "Wildcard Matching",                              "Hard",   44,   "https://leetcode.com/problems/wildcard-matching/",                                "2D DP; similar to regex"),
        (9,  "Word Break II",                                  "Hard",   140,  "https://leetcode.com/problems/word-break-ii/",                                    "Backtrack + memo — Google very frequent"),
        (10, "Burst Balloons",                                 "Hard",   312,  "https://leetcode.com/problems/burst-balloons/",                                   "Interval DP; reverse thinking"),
        (11, "Longest Increasing Path in Matrix",              "Hard",   329,  "https://leetcode.com/problems/longest-increasing-path-in-a-matrix/",             "DFS + memo (topo-sort order)"),
        (12, "Minimum Number of Refueling Stops",             "Hard",   871,  "https://leetcode.com/problems/minimum-number-of-refueling-stops/",               "Greedy + max-heap OR DP — Google frequent"),
        (13, "Stickers to Spell Word",                        "Hard",   691,  "https://leetcode.com/problems/stickers-to-spell-word/",                          "BFS on bitmask state — Google classic"),
        (14, "Distinct Subsequences",                         "Hard",   115,  "https://leetcode.com/problems/distinct-subsequences/",                           "2D DP; count ways"),
        (15, "Best Time to Buy and Sell Stock with Cooldown", "Medium", 309,  "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-with-cooldown/",  "State machine DP"),
        (16, "Decode Ways",                                   "Medium", 91,   "https://leetcode.com/problems/decode-ways/",                                      "1D DP; handle '0' edge cases"),
        (17, "Unique Paths",                                  "Medium", 62,   "https://leetcode.com/problems/unique-paths/",                                     "2D DP or combinatorics"),
        (18, "Minimum Path Sum",                              "Medium", 64,   "https://leetcode.com/problems/minimum-path-sum/",                                 "2D DP in-place"),
        (19, "Coin Change II",                                "Medium", 518,  "https://leetcode.com/problems/coin-change-ii/",                                   "Unbounded knapsack; count"),
        (20, "Partition Equal Subset Sum",                    "Medium", 416,  "https://leetcode.com/problems/partition-equal-subset-sum/",                     "0/1 knapsack; target = sum/2"),
        (21, "Maximum Rectangle (Maximal Rectangle)",         "Hard",   85,   "https://leetcode.com/problems/maximal-rectangle/",                               "Histogram DP per row"),
        (22, "Maximal Square",                                "Medium", 221,  "https://leetcode.com/problems/maximal-square/",                                   "DP; min of 3 neighbours + 1"),
        (23, "Target Sum",                                    "Medium", 494,  "https://leetcode.com/problems/target-sum/",                                       "DFS/DP or subset knapsack"),
        (24, "Interleaving String",                           "Medium", 97,   "https://leetcode.com/problems/interleaving-string/",                              "2D DP"),
        (25, "Minimum Difficulty of a Job Schedule",          "Hard",   1335, "https://leetcode.com/problems/minimum-difficulty-of-a-job-schedule/",            "2D DP; d days × n jobs"),
        (26, "Stone Game",                                    "Medium", 877,  "https://leetcode.com/problems/stone-game/",                                       "Game theory DP or math insight"),
        (27, "Best Time Buy Sell Stock III",                  "Hard",   123,  "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-iii/",             "State machine; at most 2 txns"),
        (28, "Best Time Buy Sell Stock IV",                   "Hard",   188,  "https://leetcode.com/problems/best-time-to-buy-and-sell-stock-iv/",              "State machine; at most k txns"),
    ]
)

# ── 9. Backtracking ───────────────────────────────────────────────────────────
categories["Backtracking"] = (
    "Google uses backtracking to test constraint propagation and pruning. Always identify: choices, constraints, and goals.",
    [
        (1,  "Subsets",                                  "Medium", 78,  "https://leetcode.com/problems/subsets/",                                         "Include/exclude; 2^n"),
        (2,  "Subsets II",                               "Medium", 90,  "https://leetcode.com/problems/subsets-ii/",                                      "Sort + skip duplicates"),
        (3,  "Permutations",                             "Medium", 46,  "https://leetcode.com/problems/permutations/",                                     "Swap-based or used[] array"),
        (4,  "Combination Sum",                          "Medium", 39,  "https://leetcode.com/problems/combination-sum/",                                  "Backtrack; reuse allowed"),
        (5,  "Combination Sum II",                       "Medium", 40,  "https://leetcode.com/problems/combination-sum-ii/",                               "Sort + skip used duplicates"),
        (6,  "N-Queens",                                 "Hard",   51,  "https://leetcode.com/problems/n-queens/",                                         "Row-by-row; col+diag sets — Google classic"),
        (7,  "N-Queens II",                              "Hard",   52,  "https://leetcode.com/problems/n-queens-ii/",                                      "Count only; same approach"),
        (8,  "Sudoku Solver",                            "Hard",   37,  "https://leetcode.com/problems/sudoku-solver/",                                    "Backtrack cell-by-cell; constraint sets"),
        (9,  "Word Search",                              "Medium", 79,  "https://leetcode.com/problems/word-search/",                                      "DFS + visited mark"),
        (10, "Word Search II",                           "Hard",   212, "https://leetcode.com/problems/word-search-ii/",                                   "Trie + DFS backtrack"),
        (11, "Palindrome Partitioning",                 "Medium", 131, "https://leetcode.com/problems/palindrome-partitioning/",                          "Backtrack + is-palindrome check"),
        (12, "Expression Add Operators",                "Hard",   282, "https://leetcode.com/problems/expression-add-operators/",                         "Backtrack with prev-term for * — Google frequent"),
        (13, "Remove Invalid Parentheses",              "Hard",   301, "https://leetcode.com/problems/remove-invalid-parentheses/",                      "BFS level / DFS with pruning"),
        (14, "Letter Combinations of a Phone Number",   "Medium", 17,  "https://leetcode.com/problems/letter-combinations-of-a-phone-number/",           "Backtrack; map digits to chars"),
    ]
)

# ── 10. Heap, Greedy & Intervals ──────────────────────────────────────────────
categories["Heap, Greedy & Intervals"] = (
    "Google tests heap-based k-selection, scheduling, and interval sweep. For greedy, always prove the local choice is globally optimal.",
    [
        (1,  "Find Median from Data Stream",             "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",                   "Two heaps (max+min)"),
        (2,  "Merge K Sorted Lists",                     "Hard",   23,   "https://leetcode.com/problems/merge-k-sorted-lists/",                           "Min-heap of k pointers"),
        (3,  "K Closest Points to Origin",               "Medium", 973,  "https://leetcode.com/problems/k-closest-points-to-origin/",                    "Max-heap or quickselect"),
        (4,  "Top K Frequent Elements",                  "Medium", 347,  "https://leetcode.com/problems/top-k-frequent-elements/",                       "Min-heap of size k"),
        (5,  "Kth Largest Element in an Array",          "Medium", 215,  "https://leetcode.com/problems/kth-largest-element-in-an-array/",               "Min-heap or quickselect"),
        (6,  "Task Scheduler",                           "Medium", 621,  "https://leetcode.com/problems/task-scheduler/",                                  "Greedy; fill cooling slots"),
        (7,  "IPO (Maximize Capital)",                   "Hard",   502,  "https://leetcode.com/problems/ipo/",                                             "Two heaps; sort by capital then pick max profit"),
        (8,  "Minimum Number of Refueling Stops",       "Hard",   871,  "https://leetcode.com/problems/minimum-number-of-refueling-stops/",              "Greedy + max-heap — Google frequent"),
        (9,  "Merge Intervals",                          "Medium", 56,   "https://leetcode.com/problems/merge-intervals/",                                 "Sort + sweep"),
        (10, "Insert Interval",                          "Medium", 57,   "https://leetcode.com/problems/insert-interval/",                                 "Linear scan + merge"),
        (11, "Non-overlapping Intervals",               "Medium", 435,  "https://leetcode.com/problems/non-overlapping-intervals/",                      "Sort by end; greedy keep"),
        (12, "Minimum Interval to Include Each Query",  "Hard",   1851, "https://leetcode.com/problems/minimum-interval-to-include-each-query/",        "Sort + min-heap sweep"),
        (13, "Gas Station",                             "Medium", 134,  "https://leetcode.com/problems/gas-station/",                                     "Greedy; single pass"),
        (14, "Jump Game II",                            "Medium", 45,   "https://leetcode.com/problems/jump-game-ii/",                                    "Greedy; track max reach per level"),
        (15, "Maximum Profit in Job Scheduling",        "Hard",   1235, "https://leetcode.com/problems/maximum-profit-in-job-scheduling/",               "Sort by end + DP + binary search"),
    ]
)

# ── 11. Tries ─────────────────────────────────────────────────────────────────
categories["Tries"] = (
    "Google Search autocomplete, spell-checking, and IP routing all use tries. Know insert/search/startsWith and deletion.",
    [
        (1, "Implement Trie (Prefix Tree)",               "Medium", 208, "https://leetcode.com/problems/implement-trie-prefix-tree/",                      "26-child or hash-map node"),
        (2, "Design Add and Search Words Data Structure", "Medium", 211, "https://leetcode.com/problems/design-add-and-search-words-data-structure/",     "Trie + DFS for '.'"),
        (3, "Word Search II",                             "Hard",   212, "https://leetcode.com/problems/word-search-ii/",                                   "Trie + backtrack on grid"),
        (4, "Replace Words",                              "Medium", 648, "https://leetcode.com/problems/replace-words/",                                    "Trie prefix lookup"),
        (5, "Maximum XOR of Two Numbers",                 "Medium", 421, "https://leetcode.com/problems/maximum-xor-of-two-numbers-in-an-array/",         "Binary (bit) trie; greedy bit choice"),
        (6, "Implement Magic Dictionary",                 "Medium", 676, "https://leetcode.com/problems/implement-magic-dictionary/",                      "Trie + DFS with 1 allowed mismatch"),
        (7, "Concatenated Words",                         "Hard",   472, "https://leetcode.com/problems/concatenated-words/",                              "Trie + DP / DFS with memo"),
    ]
)

# ── 12. Math, Bit Manipulation & Geometry ────────────────────────────────────
categories["Math & Bit Manipulation"] = (
    "Google interviews include bit-level tricks, modular arithmetic, prime sieves, and matrix geometry. O(1) space solutions valued.",
    [
        (1,  "Single Number",                   "Easy",   136, "https://leetcode.com/problems/single-number/",                           "XOR all elements"),
        (2,  "Number of 1 Bits (Hamming Weight)","Easy",  191, "https://leetcode.com/problems/number-of-1-bits/",                       "n & (n-1) trick"),
        (3,  "Counting Bits",                   "Easy",   338, "https://leetcode.com/problems/counting-bits/",                          "DP: dp[i] = dp[i>>1] + (i&1)"),
        (4,  "Reverse Bits",                    "Easy",   190, "https://leetcode.com/problems/reverse-bits/",                           "Bit-by-bit or divide & cache"),
        (5,  "Missing Number",                  "Easy",   268, "https://leetcode.com/problems/missing-number/",                         "XOR or Gauss sum"),
        (6,  "Sum of Two Integers (No +/-)",    "Medium", 371, "https://leetcode.com/problems/sum-of-two-integers/",                   "Carry with AND, XOR loop"),
        (7,  "Maximum XOR of Two Numbers",      "Medium", 421, "https://leetcode.com/problems/maximum-xor-of-two-numbers-in-an-array/","Bit trie; greedy"),
        (8,  "Single Number II",                "Medium", 137, "https://leetcode.com/problems/single-number-ii/",                       "Count bits mod 3"),
        (9,  "Pow(x, n)",                       "Medium", 50,  "https://leetcode.com/problems/powx-n/",                                 "Fast exponentiation; handle negatives"),
        (10, "Spiral Matrix",                   "Medium", 54,  "https://leetcode.com/problems/spiral-matrix/",                          "Layer-by-layer or direction array"),
        (11, "Rotate Image",                    "Medium", 48,  "https://leetcode.com/problems/rotate-image/",                           "Transpose then reverse rows"),
        (12, "Set Matrix Zeroes",               "Medium", 73,  "https://leetcode.com/problems/set-matrix-zeroes/",                     "Use first row/col as flags"),
        (13, "Count Primes",                    "Medium", 204, "https://leetcode.com/problems/count-primes/",                           "Sieve of Eratosthenes"),
        (14, "Max Points on a Line",            "Hard",   149, "https://leetcode.com/problems/max-points-on-a-line/",                   "Slope as fraction (gcd); map per anchor — Google classic"),
        (15, "Multiply Strings",               "Medium", 43,  "https://leetcode.com/problems/multiply-strings/",                      "Grade-school multiplication"),
        (16, "Integer to English Words",        "Hard",   273, "https://leetcode.com/problems/integer-to-english-words/",              "Recursive by chunk-of-3 — Google frequent"),
        (17, "Happy Number",                    "Easy",   202, "https://leetcode.com/problems/happy-number/",                          "Floyd cycle or seen set"),
        (18, "Reverse Integer",                 "Medium", 7,   "https://leetcode.com/problems/reverse-integer/",                       "Overflow check with 32-bit bounds"),
    ]
)

# ── 13. Design / System Design ────────────────────────────────────────────────
categories["Design"] = (
    "Google design questions test OOP, scalable data structures, and trade-off analysis. Explain time/space complexity clearly.",
    [
        (1,  "LRU Cache",                              "Medium", 146,  "https://leetcode.com/problems/lru-cache/",                                   "Doubly-LL + hash map; O(1) ops"),
        (2,  "LFU Cache",                              "Hard",   460,  "https://leetcode.com/problems/lfu-cache/",                                   "Two maps + DLL; O(1) ops"),
        (3,  "Serialize and Deserialize Binary Tree",  "Hard",   297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/",      "BFS or preorder string encoding"),
        (4,  "Design Twitter",                         "Medium", 355,  "https://leetcode.com/problems/design-twitter/",                              "Heap-merge recent tweets"),
        (5,  "Find Median from Data Stream",           "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",               "Two heaps"),
        (6,  "Range Sum Query — Mutable",              "Medium", 307,  "https://leetcode.com/problems/range-sum-query-mutable/",                    "Segment tree or BIT (Fenwick)"),
        (7,  "Design Add and Search Words",            "Medium", 211,  "https://leetcode.com/problems/design-add-and-search-words-data-structure/", "Trie + wildcard DFS"),
        (8,  "Snapshot Array",                         "Medium", 1146, "https://leetcode.com/problems/snapshot-array/",                             "Binary search on snap history"),
        (9,  "My Calendar I",                          "Medium", 729,  "https://leetcode.com/problems/my-calendar-i/",                              "Sorted list / BST of intervals"),
        (10, "Design Underground System",              "Medium", 1396, "https://leetcode.com/problems/design-underground-system/",                  "Two hash maps; average running total"),
        (11, "Implement Trie",                         "Medium", 208,  "https://leetcode.com/problems/implement-trie-prefix-tree/",                 "Array or map children"),
        (12, "Design HashMap",                         "Easy",   706,  "https://leetcode.com/problems/design-hashmap/",                             "Chaining or open addressing"),
        (13, "Flatten Nested List Iterator",           "Medium", 341,  "https://leetcode.com/problems/flatten-nested-list-iterator/",              "Stack of iterators"),
        (14, "Maximum Profit in Job Scheduling",       "Hard",   1235, "https://leetcode.com/problems/maximum-profit-in-job-scheduling/",          "Sort + DP + binary search"),
    ]
)

# ── 14. Google — High Frequency Master List ───────────────────────────────────
categories["Google High Frequency"] = (
    "These problems appear most often in reported Google interviews (via Glassdoor, Blind, LeetCode Discuss). Solve all before your interview.",
    [
        (1,  "Median of Two Sorted Arrays",                 "Hard",   4,    "https://leetcode.com/problems/median-of-two-sorted-arrays/",                     "Binary search on partition"),
        (2,  "Regular Expression Matching",                 "Hard",   10,   "https://leetcode.com/problems/regular-expression-matching/",                     "2D DP"),
        (3,  "Text Justification",                          "Hard",   68,   "https://leetcode.com/problems/text-justification/",                               "String simulation"),
        (4,  "Minimum Window Substring",                    "Hard",   76,   "https://leetcode.com/problems/minimum-window-substring/",                        "Sliding window"),
        (5,  "Largest Rectangle in Histogram",              "Hard",   84,   "https://leetcode.com/problems/largest-rectangle-in-histogram/",                  "Monotonic stack"),
        (6,  "Word Break II",                               "Hard",   140,  "https://leetcode.com/problems/word-break-ii/",                                    "Backtrack + memo"),
        (7,  "Accounts Merge",                              "Medium", 721,  "https://leetcode.com/problems/accounts-merge/",                                   "Union-Find / DFS"),
        (8,  "N-Queens",                                    "Hard",   51,   "https://leetcode.com/problems/n-queens/",                                         "Backtracking"),
        (9,  "Wildcard Matching",                           "Hard",   44,   "https://leetcode.com/problems/wildcard-matching/",                                "2D DP"),
        (10, "Serialize and Deserialize Binary Tree",       "Hard",   297,  "https://leetcode.com/problems/serialize-and-deserialize-binary-tree/",           "BFS or preorder"),
        (11, "Longest Increasing Path in Matrix",           "Hard",   329,  "https://leetcode.com/problems/longest-increasing-path-in-a-matrix/",            "DFS + memo"),
        (12, "Burst Balloons",                              "Hard",   312,  "https://leetcode.com/problems/burst-balloons/",                                   "Interval DP"),
        (13, "Word Ladder",                                 "Hard",   127,  "https://leetcode.com/problems/word-ladder/",                                      "BFS"),
        (14, "Swim in Rising Water",                        "Hard",   778,  "https://leetcode.com/problems/swim-in-rising-water/",                            "Dijkstra / BS+BFS"),
        (15, "Stickers to Spell Word",                     "Hard",   691,  "https://leetcode.com/problems/stickers-to-spell-word/",                          "BFS bitmask"),
        (16, "Vertical Order Traversal of Binary Tree",    "Hard",   987,  "https://leetcode.com/problems/vertical-order-traversal-of-a-binary-tree/",      "DFS with col/row keys"),
        (17, "Minimum Number of Refueling Stops",          "Hard",   871,  "https://leetcode.com/problems/minimum-number-of-refueling-stops/",               "Greedy + max-heap"),
        (18, "Max Points on a Line",                        "Hard",   149,  "https://leetcode.com/problems/max-points-on-a-line/",                            "Slope hash map"),
        (19, "Integer to English Words",                    "Hard",   273,  "https://leetcode.com/problems/integer-to-english-words/",                       "Chunk-of-3 recursion"),
        (20, "Expression Add Operators",                    "Hard",   282,  "https://leetcode.com/problems/expression-add-operators/",                        "Backtrack with prev-term"),
        (21, "Count of Smaller Numbers After Self",         "Hard",   315,  "https://leetcode.com/problems/count-of-smaller-numbers-after-self/",            "Merge sort / BIT"),
        (22, "LFU Cache",                                   "Hard",   460,  "https://leetcode.com/problems/lfu-cache/",                                        "Two maps + DLL"),
        (23, "Remove Invalid Parentheses",                  "Hard",   301,  "https://leetcode.com/problems/remove-invalid-parentheses/",                     "BFS / DFS pruned"),
        (24, "Find Median from Data Stream",                "Hard",   295,  "https://leetcode.com/problems/find-median-from-data-stream/",                   "Two heaps"),
        (25, "Reconstruct Itinerary",                       "Hard",   332,  "https://leetcode.com/problems/reconstruct-itinerary/",                          "Hierholzer Eulerian path"),
        (26, "Critical Connections in a Network",           "Hard",   1192, "https://leetcode.com/problems/critical-connections-in-a-network/",              "Tarjan's bridge"),
        (27, "Minimum Window Subsequence",                  "Hard",   727,  "https://leetcode.com/problems/minimum-window-subsequence/",                     "Two-pointer scan"),
        (28, "Minimum Difficulty of a Job Schedule",        "Hard",   1335, "https://leetcode.com/problems/minimum-difficulty-of-a-job-schedule/",           "2D DP"),
        (29, "Word Search II",                              "Hard",   212,  "https://leetcode.com/problems/word-search-ii/",                                   "Trie + backtrack"),
        (30, "Sudoku Solver",                               "Hard",   37,   "https://leetcode.com/problems/sudoku-solver/",                                    "Backtrack + constraint sets"),
    ]
)

# ═══════════════════════════════════════════════════════════════════════════════
#  BUILD CATEGORY SHEETS
# ═══════════════════════════════════════════════════════════════════════════════
for cat_name, (desc, rows) in categories.items():
    add_category_sheet(wb, cat_name, desc, rows)

# ═══════════════════════════════════════════════════════════════════════════════
#  GOOGLE 8-WEEK STUDY PLAN SHEET
# ═══════════════════════════════════════════════════════════════════════════════
ws_plan = wb.create_sheet(title="Google 8-Week Study Plan")

ws_plan.merge_cells("A1:E1")
c = ws_plan["A1"]
c.value     = "Google Interview — 8-Week Intensive Study Plan"
c.fill      = FILLS["plan"]
c.font      = Font(color="FFFFFF", bold=True, size=14)
c.alignment = CENTER
ws_plan.row_dimensions[1].height = 34

plan_headers = ["Week", "Focus Area", "Topics / Sheets", "Daily Goal", "Tip"]
plan_widths  = [8,       28,           50,                40,           52]
for col, (h, w) in enumerate(zip(plan_headers, plan_widths), 1):
    cell = ws_plan.cell(row=2, column=col, value=h)
    cell.fill      = FILLS["col_hdr"]
    cell.font      = WHITE
    cell.alignment = CENTER
    cell.border    = BOX
    ws_plan.column_dimensions[get_column_letter(col)].width = w
ws_plan.row_dimensions[2].height = 20

plan_data = [
    ("Week 1",  "Foundations",          "Arrays & Hashing, Two Pointers, Sliding Window",                    "5–6 problems/day. Aim for Easy→Medium.",   "Code every problem from scratch — no looking at hints first."),
    ("Week 2",  "Linear Structures",    "Stack & Monotonic Stack, Linked Lists, Binary Search",              "5 problems/day. Time yourself (≤30 min).", "For binary search: write lo/hi/mid template once and re-use."),
    ("Week 3",  "Trees",                "Trees & BST",                                                        "4–5 problems/day. Draw recursion trees.",  "Trace DFS/BFS on paper before coding."),
    ("Week 4",  "Graphs",               "Graphs (BFS/DFS/Union-Find/Topo Sort)",                             "4 problems/day. Build adjacency list.",    "Google loves multi-source BFS and Dijkstra variants."),
    ("Week 5",  "DP Fundamentals",      "DP — 1D patterns (coin, LIS, knapsack)",                            "3–4 problems/day. Write recurrence first.","Identify: state definition, transition, base case."),
    ("Week 6",  "DP Advanced + Tries",  "DP — 2D, Interval DP, Bitmask DP, Tries",                          "3 problems/day. Review solutions deeply.", "Regex/Wildcard/Burst Balloons are Google favourites."),
    ("Week 7",  "Backtracking + Heap",  "Backtracking, Heap/Greedy/Intervals",                               "3–4 problems/day.",                        "Prune early! Always check validity before recursing."),
    ("Week 8",  "Mock + Review",        "Google High Frequency list, Math & Bits, Design",                   "2 problems/day + 2 mock interviews.",      "Use interviewing.io or Pramp. Record yourself explaining."),
]

phase_colors = ["E3F2FD", "E8F5E9", "FFF8E1", "FCE4EC", "F3E5F5", "E0F7FA", "FBE9E7", "E8EAF6"]
for r_idx, row in enumerate(plan_data, start=3):
    fill = PatternFill("solid", fgColor=phase_colors[r_idx - 3])
    for col, val in enumerate(row, start=1):
        cell = ws_plan.cell(row=r_idx, column=col, value=val)
        cell.fill      = fill
        cell.font      = BOLD if col <= 2 else NORMAL
        cell.alignment = LEFT if col > 1 else CENTER
        cell.border    = BOX
    ws_plan.row_dimensions[r_idx].height = 30

# Resources
ws_plan.merge_cells("A12:E12")
r12 = ws_plan["A12"]
r12.value     = "Essential Google Prep Resources"
r12.fill      = FILLS["plan"]
r12.font      = Font(color="FFFFFF", bold=True, size=12)
r12.alignment = CENTER
ws_plan.row_dimensions[12].height = 26

resources = [
    ("LeetCode Google Tag",          "https://leetcode.com/problemset/?search=google",                         "Filter by Google company tag — most relevant problems"),
    ("NeetCode Roadmap",             "https://neetcode.io/roadmap",                                            "Structured topic-by-topic practice with video explanations"),
    ("Blind 75",                     "https://leetcode.com/discuss/general-discussion/460599/blind-75-leetcode-questions","The classic 75-problem must-solve list"),
    ("Grind 75",                     "https://www.techinterviewhandbook.org/grind75",                          "Time-boxed version of Blind 75"),
    ("Google Interview Warmup",      "https://grow.google/certificates/interview-warmup/",                    "Google's own behavioural practice tool"),
    ("System Design Primer",         "https://github.com/donnemartin/system-design-primer",                   "For the system design round (Round 3)"),
    ("interviewing.io",              "https://interviewing.io",                                                "Anonymous mock interviews with real engineers"),
    ("Pramp",                        "https://www.pramp.com",                                                  "Peer-to-peer mock interview platform"),
    ("Designing Data-Intensive Apps","https://dataintensive.net",                                              "Best system design book (Kleppmann)"),
]

res_hdrs  = ["Resource", "Link", "Why It Helps"]
res_widths = [30, 48, 45]
for col, (h, w) in enumerate(zip(res_hdrs, res_widths), 1):
    cell = ws_plan.cell(row=13, column=col, value=h)
    cell.fill      = FILLS["col_hdr"]
    cell.font      = WHITE
    cell.alignment = CENTER
    cell.border    = BOX
    ws_plan.column_dimensions[get_column_letter(col)].width = w
ws_plan.row_dimensions[13].height = 20

# Add an unused column E for row 13
ws_plan.merge_cells("D13:E13")
ws_plan.cell(row=13, column=4, value="").fill = FILLS["col_hdr"]

for r_idx, (name, url, why) in enumerate(resources, start=14):
    fill = PatternFill("solid", fgColor="E8EAF6" if r_idx % 2 == 0 else "FFFFFF")
    ws_plan.cell(row=r_idx, column=1, value=name).fill      = fill
    ws_plan.cell(row=r_idx, column=1).font      = BOLD
    ws_plan.cell(row=r_idx, column=1).alignment = LEFT
    ws_plan.cell(row=r_idx, column=1).border    = BOX

    lc = ws_plan.cell(row=r_idx, column=2, value=url)
    lc.hyperlink  = url
    lc.font       = LINK
    lc.fill       = fill
    lc.alignment  = LEFT
    lc.border     = BOX

    ws_plan.cell(row=r_idx, column=3, value=why).fill      = fill
    ws_plan.cell(row=r_idx, column=3).font      = Font(size=9, italic=True, color="444444")
    ws_plan.cell(row=r_idx, column=3).alignment = LEFT
    ws_plan.cell(row=r_idx, column=3).border    = BOX

    ws_plan.merge_cells(f"D{r_idx}:E{r_idx}")
    ws_plan.cell(row=r_idx, column=4, value="").fill   = fill
    ws_plan.cell(row=r_idx, column=4).border = BOX
    ws_plan.row_dimensions[r_idx].height = 18

ws_plan.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════════════
#  OVERVIEW / INDEX SHEET (inserted first)
# ═══════════════════════════════════════════════════════════════════════════════
ws_idx = wb.create_sheet(title="Overview", index=0)

ws_idx.merge_cells("A1:D1")
c = ws_idx["A1"]
c.value     = "Google DSA Interview Prep — Overview"
c.fill      = FILLS["title"]
c.font      = Font(color="FFFFFF", bold=True, size=14)
c.alignment = CENTER
ws_idx.row_dimensions[1].height = 34

# Google colour stripe
for col, color in enumerate([GOOGLE_BLUE, GOOGLE_RED, GOOGLE_YELLOW, GOOGLE_GREEN], 1):
    cell = ws_idx.cell(row=2, column=col, value="")
    cell.fill = PatternFill("solid", fgColor=color)
ws_idx.row_dimensions[2].height = 6

# Tips
tips = [
    "Think out loud — Google interviewers value your reasoning process as much as the correct answer.",
    "Start with brute force, then optimise. Never jump to the optimal solution without explaining the path.",
    "Always clarify constraints: input size, duplicates, negative numbers, empty inputs.",
    "Discuss trade-offs: time vs space, recursion vs iteration, clarity vs performance.",
    "Google values 'Googliness': intellectual humility, curiosity, and collaborative problem-solving.",
]
ws_idx.merge_cells("A3:D3")
ws_idx.cell(row=3, column=1, value="Google Interview Tips").fill = FILLS["col_hdr"]
ws_idx.cell(row=3, column=1).font      = WHITE
ws_idx.cell(row=3, column=1).alignment = CENTER
ws_idx.cell(row=3, column=1).border    = BOX
ws_idx.row_dimensions[3].height = 20

for r_idx, tip in enumerate(tips, start=4):
    ws_idx.merge_cells(f"A{r_idx}:D{r_idx}")
    cell = ws_idx.cell(row=r_idx, column=1, value=f"  • {tip}")
    cell.fill      = FILLS["tip"]
    cell.font      = Font(size=10, italic=True, color="333333")
    cell.alignment = LEFT
    cell.border    = BOX
    ws_idx.row_dimensions[r_idx].height = 22

# Table header
tbl_row = len(tips) + 5
ws_idx.merge_cells(f"A{tbl_row}:D{tbl_row}")
ws_idx.cell(row=tbl_row, column=1, value="Sheet Directory").fill  = FILLS["col_hdr"]
ws_idx.cell(row=tbl_row, column=1).font      = WHITE
ws_idx.cell(row=tbl_row, column=1).alignment = CENTER
ws_idx.cell(row=tbl_row, column=1).border    = BOX
ws_idx.row_dimensions[tbl_row].height = 22

col_hdrs   = ["Category / Sheet", "# Questions", "Difficulty Mix", "Google Focus Level"]
col_widths = [38, 14, 28, 28]
for col, (h, w) in enumerate(zip(col_hdrs, col_widths), 1):
    cell = ws_idx.cell(row=tbl_row + 1, column=col, value=h)
    cell.fill      = PatternFill("solid", fgColor="42A5F5")
    cell.font      = WHITE
    cell.alignment = CENTER
    cell.border    = BOX
    ws_idx.column_dimensions[get_column_letter(col)].width = w
ws_idx.row_dimensions[tbl_row + 1].height = 20

# Difficulty summaries
def diff_summary(rows):
    e = sum(1 for r in rows if r[2] == "Easy")
    m = sum(1 for r in rows if r[2] == "Medium")
    h = sum(1 for r in rows if r[2] == "Hard")
    return f"E:{e}  M:{m}  H:{h}"

focus_map = {
    "Arrays & Hashing"           : "★★★★☆  Core",
    "Strings & Sliding Window"   : "★★★★★  Very High",
    "Two Pointers & Binary Search": "★★★★☆  High",
    "Stack & Monotonic Stack"    : "★★★☆☆  Medium",
    "Linked Lists"               : "★★★☆☆  Medium",
    "Trees & BST"                : "★★★★★  Very High",
    "Graphs"                     : "★★★★★  Very High",
    "Dynamic Programming"        : "★★★★★  Highest",
    "Backtracking"               : "★★★★☆  High",
    "Heap, Greedy & Intervals"   : "★★★★☆  High",
    "Tries"                      : "★★★☆☆  Medium",
    "Math & Bit Manipulation"    : "★★★☆☆  Medium",
    "Design"                     : "★★★★☆  High",
    "Google High Frequency"      : "★★★★★  Must Do",
    "Google 8-Week Study Plan"   : "—  Study Guide",
}

ALT1 = PatternFill("solid", fgColor="E3F2FD")
ALT2 = PatternFill("solid", fgColor="FFFFFF")
MUST = PatternFill("solid", fgColor="FFCDD2")

for r_idx, (cat_name, (_, rows)) in enumerate(categories.items(), start=tbl_row + 2):
    fill = MUST if cat_name == "Google High Frequency" else (ALT1 if r_idx % 2 == 0 else ALT2)

    ws_idx.cell(row=r_idx, column=1, value=cat_name).fill      = fill
    ws_idx.cell(row=r_idx, column=1).font      = BOLD
    ws_idx.cell(row=r_idx, column=1).alignment = LEFT
    ws_idx.cell(row=r_idx, column=1).border    = BOX

    ws_idx.cell(row=r_idx, column=2, value=len(rows)).fill      = fill
    ws_idx.cell(row=r_idx, column=2).font      = NORMAL
    ws_idx.cell(row=r_idx, column=2).alignment = CENTER
    ws_idx.cell(row=r_idx, column=2).border    = BOX

    ws_idx.cell(row=r_idx, column=3, value=diff_summary(rows)).fill      = fill
    ws_idx.cell(row=r_idx, column=3).font      = Font(size=9, color="444444")
    ws_idx.cell(row=r_idx, column=3).alignment = CENTER
    ws_idx.cell(row=r_idx, column=3).border    = BOX

    ws_idx.cell(row=r_idx, column=4, value=focus_map.get(cat_name, "")).fill      = fill
    ws_idx.cell(row=r_idx, column=4).font      = Font(size=10, bold=(cat_name == "Google High Frequency"), color="C62828" if "Highest" in focus_map.get(cat_name,"") or "Must" in focus_map.get(cat_name,"") else "333333")
    ws_idx.cell(row=r_idx, column=4).alignment = LEFT
    ws_idx.cell(row=r_idx, column=4).border    = BOX
    ws_idx.row_dimensions[r_idx].height = 18

# Study plan row
last_r = tbl_row + 2 + len(categories)
ws_idx.cell(row=last_r, column=1, value="Google 8-Week Study Plan").fill      = PatternFill("solid", fgColor="C8E6C9")
ws_idx.cell(row=last_r, column=1).font      = BOLD
ws_idx.cell(row=last_r, column=1).alignment = LEFT
ws_idx.cell(row=last_r, column=1).border    = BOX
ws_idx.cell(row=last_r, column=2, value="—").alignment  = CENTER
ws_idx.cell(row=last_r, column=2).border    = BOX
ws_idx.cell(row=last_r, column=2).fill      = PatternFill("solid", fgColor="C8E6C9")
ws_idx.cell(row=last_r, column=3, value="—").alignment  = CENTER
ws_idx.cell(row=last_r, column=3).border    = BOX
ws_idx.cell(row=last_r, column=3).fill      = PatternFill("solid", fgColor="C8E6C9")
ws_idx.cell(row=last_r, column=4, value="8-week prep roadmap + resources").fill = PatternFill("solid", fgColor="C8E6C9")
ws_idx.cell(row=last_r, column=4).font      = Font(size=9, italic=True, color="444444")
ws_idx.cell(row=last_r, column=4).alignment = LEFT
ws_idx.cell(row=last_r, column=4).border    = BOX
ws_idx.row_dimensions[last_r].height = 18

# Total
total_q = sum(len(rows) for _, rows in categories.values())
ws_idx.merge_cells(f"A{last_r+1}:D{last_r+1}")
total_cell = ws_idx.cell(row=last_r + 1, column=1, value=f"Total: {total_q} problems across {len(categories)} categories")
total_cell.fill      = PatternFill("solid", fgColor="1A237E")
total_cell.font      = Font(color="FFFFFF", bold=True, size=11)
total_cell.alignment = CENTER
total_cell.border    = BOX
ws_idx.row_dimensions[last_r + 1].height = 22

ws_idx.freeze_panes = f"A{tbl_row + 2}"

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = r"c:\Users\MaitySom\Desktop\Somneel\Somneel_Maity\K_Tech\DSA\Google_DSA_Interview_Prep.xlsx"
wb.save(out_path)
print(f"✅ Saved: {out_path}")
print(f"   Sheets : {len(wb.sheetnames)}")
print(f"   Total questions : {total_q}")
for cat_name, (_, rows) in categories.items():
    print(f"   {cat_name:<38} {len(rows):>3} questions")
