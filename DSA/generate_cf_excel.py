import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── Palette ────────────────────────────────────────────────────────────────────
TITLE_FILL   = PatternFill("solid", fgColor="B71C1C")   # deep red (CF brand)
COLHDR_FILL  = PatternFill("solid", fgColor="C62828")   # red
OVERVIEW_HDR = PatternFill("solid", fgColor="4E342E")   # brown
PLAN_FILL    = PatternFill("solid", fgColor="1B5E20")   # dark green

DIFF_FILL = {
    "Beginner": PatternFill("solid", fgColor="E8F5E9"),  # light green
    "Easy":     PatternFill("solid", fgColor="FFF9C4"),  # light yellow
    "Medium":   PatternFill("solid", fgColor="FFE0B2"),  # light orange
    "Hard":     PatternFill("solid", fgColor="FFCDD2"),  # light red
    "Expert":   PatternFill("solid", fgColor="E1BEE7"),  # light purple
}
DIFF_DOT = {
    "Beginner": "●",
    "Easy":     "●",
    "Medium":   "●",
    "Hard":     "●",
    "Expert":   "●",
}

WHITE  = Font(color="FFFFFF", bold=True, size=11)
BOLD   = Font(bold=True, size=10)
NORMAL = Font(size=10)
LINK   = Font(color="1565C0", underline="single", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN   = Side(style="thin", color="BDBDBD")
BOX    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def safe(name):
    for ch in r"/\?*[]:'":
        name = name.replace(ch, "-")
    return name[:31]


def add_topic_sheet(wb, sheet_name, rows):
    """rows: (no, name, difficulty, cf_rating, url)"""
    ws = wb.create_sheet(title=safe(sheet_name))

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = sheet_name
    c.fill      = TITLE_FILL
    c.font      = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    hdrs   = ["#", "Problem Name", "Difficulty", "CF Rating", "Contest", "Problem Link"]
    widths = [5, 50, 12, 12, 14, 52]
    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        cell            = ws.cell(row=2, column=col, value=h)
        cell.fill       = COLHDR_FILL
        cell.font       = WHITE
        cell.alignment  = CENTER
        cell.border     = BOX
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    for r, (no, name, diff, rating, url) in enumerate(rows, start=3):
        fill = DIFF_FILL.get(diff, PatternFill())
        contest = url.rstrip("/").rsplit("/", 2)
        # extract "CF XXXX/Y" from URL
        parts = url.replace("https://codeforces.com/problemset/problem/", "").rstrip("/").split("/")
        contest_label = f"{parts[0]}/{parts[1]}" if len(parts) >= 2 else ""

        ws.cell(row=r, column=1, value=no).alignment       = CENTER
        ws.cell(row=r, column=2, value=name).alignment     = LEFT
        ws.cell(row=r, column=3, value=diff).alignment     = CENTER
        ws.cell(row=r, column=4, value=rating).alignment   = CENTER
        ws.cell(row=r, column=5, value=contest_label).alignment = CENTER

        lc            = ws.cell(row=r, column=6, value=f"Open Problem →")
        lc.hyperlink  = url
        lc.font       = LINK
        lc.alignment  = LEFT

        for col in range(1, 7):
            cell        = ws.cell(row=r, column=col)
            cell.fill   = fill
            cell.border = BOX
            if col not in (3, 6):
                cell.font = NORMAL
            if col == 3:
                cell.font = Font(bold=True, size=10,
                                 color={"Beginner": "2E7D32",
                                        "Easy":     "F57F17",
                                        "Medium":   "E65100",
                                        "Hard":     "B71C1C",
                                        "Expert":   "6A1B9A"}.get(diff, "000000"))
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A3"
    return ws


# ══════════════════════════════════════════════════════════════════════════════
#  DATA  —  (no, name, difficulty, cf_rating, url)
#
#  Difficulty bands:
#    Beginner  800–1000
#    Easy     1100–1400
#    Medium   1500–1800
#    Hard     1900–2200
#    Expert   2300+
# ══════════════════════════════════════════════════════════════════════════════

topics = {}

# ── 1. Implementation / Simulation ────────────────────────────────────────────
topics["Implementation"] = [
    (1,  "Watermelon",                    "Beginner", 800,  "https://codeforces.com/problemset/problem/4/A"),
    (2,  "Theatre Square",                "Beginner", 800,  "https://codeforces.com/problemset/problem/1/A"),
    (3,  "Next Round",                    "Beginner", 800,  "https://codeforces.com/problemset/problem/158/A"),
    (4,  "Team",                          "Beginner", 800,  "https://codeforces.com/problemset/problem/231/A"),
    (5,  "Stones on the Table",           "Beginner", 800,  "https://codeforces.com/problemset/problem/266/A"),
    (6,  "Helpful Maths",                 "Beginner", 800,  "https://codeforces.com/problemset/problem/339/A"),
    (7,  "Domino Piling",                 "Beginner", 800,  "https://codeforces.com/problemset/problem/50/A"),
    (8,  "Boy or Girl",                   "Beginner", 800,  "https://codeforces.com/problemset/problem/236/A"),
    (9,  "Sum of Round Numbers",          "Beginner", 800,  "https://codeforces.com/problemset/problem/1352/A"),
    (10, "Maximum GCD",                   "Beginner", 800,  "https://codeforces.com/problemset/problem/1370/A"),
    (11, "Deadline",                      "Beginner", 900,  "https://codeforces.com/problemset/problem/1288/A"),
    (12, "Collecting Packages",           "Beginner", 900,  "https://codeforces.com/problemset/problem/1294/B"),
    (13, "Three Swimmers",                "Beginner", 1000, "https://codeforces.com/problemset/problem/1492/A"),
    (14, "I Wanna Be the Guy",            "Easy",     1100, "https://codeforces.com/problemset/problem/469/A"),
    (15, "Beautiful Matrix",              "Easy",     1400, "https://codeforces.com/problemset/problem/263/A"),
    (16, "Berland Crossword",             "Easy",     1400, "https://codeforces.com/problemset/problem/1494/B"),
    (17, "Young Physicist",               "Beginner", 1000, "https://codeforces.com/problemset/problem/69/A"),
    (18, "Lucky Division",                "Easy",     1100, "https://codeforces.com/problemset/problem/121/C"),
]

# ── 2. Greedy ─────────────────────────────────────────────────────────────────
topics["Greedy"] = [
    (1,  "Remove Smallest",               "Beginner", 900,  "https://codeforces.com/problemset/problem/1399/A"),
    (2,  "Gifts Fixing",                  "Easy",     1100, "https://codeforces.com/problemset/problem/1399/B"),
    (3,  "Boats Competition",             "Easy",     1200, "https://codeforces.com/problemset/problem/1399/C"),
    (4,  "Dominant Piranha",              "Easy",     1200, "https://codeforces.com/problemset/problem/1433/C"),
    (5,  "Different Divisors",            "Easy",     1200, "https://codeforces.com/problemset/problem/1474/B"),
    (6,  "Pokémon Quest",                 "Easy",     1200, "https://codeforces.com/problemset/problem/1420/C"),
    (7,  "Multiply by 2, divide by 6",    "Easy",     1300, "https://codeforces.com/problemset/problem/1374/B"),
    (8,  "Ternary XOR",                   "Easy",     1400, "https://codeforces.com/problemset/problem/1328/C"),
    (9,  "Add to Neighbour and Remove",   "Easy",     1400, "https://codeforces.com/problemset/problem/1462/D"),
    (10, "Make It Good",                  "Easy",     1200, "https://codeforces.com/problemset/problem/1385/C"),
    (11, "Number Game",                   "Medium",   1600, "https://codeforces.com/problemset/problem/1370/C"),
    (12, "Perfect Keyboard",              "Medium",   1600, "https://codeforces.com/problemset/problem/1303/C"),
    (13, "Replace by MEX",                "Medium",   1600, "https://codeforces.com/problemset/problem/1375/D"),
    (14, "Magic Grid",                    "Medium",   1700, "https://codeforces.com/problemset/problem/1208/C"),
    (15, "Row GCD",                       "Medium",   1800, "https://codeforces.com/problemset/problem/1458/C"),
    (16, "Ehab and Another xor Problem",  "Medium",   1700, "https://codeforces.com/problemset/problem/1174/C"),
    (17, "Binary String Prefix",          "Easy",     1300, "https://codeforces.com/problemset/problem/1367/C"),
    (18, "GCD Compression",               "Medium",   1500, "https://codeforces.com/problemset/problem/1370/B"),
]

# ── 3. Sorting ────────────────────────────────────────────────────────────────
topics["Sorting"] = [
    (1,  "Three Pairwise Maximums",       "Beginner", 800,  "https://codeforces.com/problemset/problem/1385/A"),
    (2,  "Number of Apartments",          "Beginner", 800,  "https://codeforces.com/problemset/problem/1430/A"),
    (3,  "Unique Bid Auction",            "Beginner", 800,  "https://codeforces.com/problemset/problem/1454/B"),
    (4,  "Even Array",                    "Beginner", 900,  "https://codeforces.com/problemset/problem/1367/B"),
    (5,  "Silly Mistake",                 "Beginner", 1000, "https://codeforces.com/problemset/problem/1253/B"),
    (6,  "Orac and Medians",              "Medium",   1800, "https://codeforces.com/problemset/problem/1349/B"),
    (7,  "Unique Number",                 "Easy",     1300, "https://codeforces.com/problemset/problem/1462/C"),
    (8,  "Card Game (Codeforces)",        "Beginner", 900,  "https://codeforces.com/problemset/problem/1503/B"),
    (9,  "Three from One",                "Beginner", 800,  "https://codeforces.com/problemset/problem/1509/A"),
    (10, "Balanced Array",                "Easy",     1100, "https://codeforces.com/problemset/problem/1638/B"),
    (11, "Count the Arrays",              "Medium",   1800, "https://codeforces.com/problemset/problem/1362/D"),
    (12, "Tanya and Postcard",            "Easy",     1200, "https://codeforces.com/problemset/problem/1607/B"),
    (13, "Perform Easily",                "Medium",   1800, "https://codeforces.com/problemset/problem/1529/C"),
    (14, "Minimum Ties",                  "Medium",   1600, "https://codeforces.com/problemset/problem/1487/C"),
]

# ── 4. Binary Search ──────────────────────────────────────────────────────────
topics["Binary Search"] = [
    (1,  "K-th Not Divisible by n",       "Easy",     1200, "https://codeforces.com/problemset/problem/1352/C"),
    (2,  "Bad Ugly Numbers",              "Medium",   1500, "https://codeforces.com/problemset/problem/1354/C"),
    (3,  "Corrupted Array",               "Medium",   1600, "https://codeforces.com/problemset/problem/1512/D"),
    (4,  "Maximum Median",                "Medium",   1700, "https://codeforces.com/problemset/problem/1492/C"),
    (5,  "A Strange Letter",              "Medium",   1600, "https://codeforces.com/problemset/problem/1365/D"),
    (6,  "Minimum Product",               "Medium",   1600, "https://codeforces.com/problemset/problem/1409/C"),
    (7,  "Binary Search Implementation",  "Beginner", 800,  "https://codeforces.com/problemset/problem/702/B"),
    (8,  "Ternary Search (Convex)",        "Medium",   1700, "https://codeforces.com/problemset/problem/578/C"),
    (9,  "Guess the Array",               "Easy",     1300, "https://codeforces.com/problemset/problem/1363/D"),
    (10, "Xenia and Tree",                "Hard",     2400, "https://codeforces.com/problemset/problem/342/E"),
    (11, "Parallel Binary Search",         "Hard",     2200, "https://codeforces.com/problemset/problem/1093/G"),
    (12, "The Fair Nut and Rectangles",    "Hard",     2200, "https://codeforces.com/problemset/problem/1083/E"),
    (13, "Kth Smallest Sum",              "Medium",   1700, "https://codeforces.com/problemset/problem/1399/E"),
    (14, "Searching for Sums",            "Medium",   1500, "https://codeforces.com/problemset/problem/1398/C"),
]

# ── 5. Two Pointers ───────────────────────────────────────────────────────────
topics["Two Pointers"] = [
    (1,  "Balanced Playlist",             "Medium",   1700, "https://codeforces.com/problemset/problem/1237/D"),
    (2,  "Alice, Bob and Chocolate",      "Medium",   1700, "https://codeforces.com/problemset/problem/6/C"),
    (3,  "Binary String to Subsequences", "Medium",   1600, "https://codeforces.com/problemset/problem/1399/D"),
    (4,  "Maximum Good Subarray",         "Easy",     1400, "https://codeforces.com/problemset/problem/1367/D"),
    (5,  "Rescue Nibel!",                 "Easy",     1400, "https://codeforces.com/problemset/problem/1388/C"),
    (6,  "Captain Flint and Rectangles",  "Medium",   1600, "https://codeforces.com/problemset/problem/1148/C"),
    (7,  "Three Parts of Array",          "Beginner", 900,  "https://codeforces.com/problemset/problem/1thirds/A"),
    (8,  "Min Cost String",               "Medium",   1600, "https://codeforces.com/problemset/problem/1511/D"),
    (9,  "Consecutive Sum",               "Easy",     1300, "https://codeforces.com/problemset/problem/1506/C"),
    (10, "Good Subarrays",                "Medium",   1500, "https://codeforces.com/problemset/problem/1398/C"),
    (11, "Fixed Point Removal",           "Hard",     2200, "https://codeforces.com/problemset/problem/1439/C"),
    (12, "Palindrome Pairs",              "Medium",   1600, "https://codeforces.com/problemset/problem/1512/E"),
]

# ── 6. Strings ────────────────────────────────────────────────────────────────
topics["Strings"] = [
    (1,  "Valerii the Horse",             "Easy",     1200, "https://codeforces.com/problemset/problem/1369/B"),
    (2,  "String Equality",               "Easy",     1400, "https://codeforces.com/problemset/problem/1461/C"),
    (3,  "String Deletion",               "Medium",   1600, "https://codeforces.com/problemset/problem/1430/D"),
    (4,  "Prefix Flip",                   "Medium",   1700, "https://codeforces.com/problemset/problem/1382/C"),
    (5,  "Compress Words",                "Hard",     2000, "https://codeforces.com/problemset/problem/1200/E"),
    (6,  "Codeforces Is Terminating",     "Medium",   1600, "https://codeforces.com/problemset/problem/1517/C"),
    (7,  "Longest Common Subsequence",    "Medium",   1700, "https://codeforces.com/problemset/problem/463/D"),
    (8,  "Palindrome Counting",           "Easy",     1300, "https://codeforces.com/problemset/problem/1418/D"),
    (9,  "Good String",                   "Medium",   1500, "https://codeforces.com/problemset/problem/1389/C"),
    (10, "Adding Powers",                 "Easy",     1400, "https://codeforces.com/problemset/problem/1217/C"),
    (11, "Perform the Combo",             "Easy",     1200, "https://codeforces.com/problemset/problem/1221/C"),
    (12, "Worm Turing Machine",           "Medium",   1600, "https://codeforces.com/problemset/problem/1405/D"),
    (13, "K String",                      "Easy",     1300, "https://codeforces.com/problemset/problem/1516/C"),
    (14, "Beautiful Partition",           "Medium",   1700, "https://codeforces.com/problemset/problem/1482/C"),
]

# ── 7. Math & Number Theory ───────────────────────────────────────────────────
topics["Math & Number Theory"] = [
    (1,  "Minutes Before the New Year",   "Beginner", 800,  "https://codeforces.com/problemset/problem/1283/A"),
    (2,  "Minimal Inverse Descent",       "Easy",     1200, "https://codeforces.com/problemset/problem/1474/B"),
    (3,  "Modular Exponentiation",        "Easy",     1100, "https://codeforces.com/problemset/problem/913/B"),
    (4,  "New Year's Number",             "Beginner", 900,  "https://codeforces.com/problemset/problem/1481/B"),
    (5,  "Count the Multiples of 2",      "Easy",     1200, "https://codeforces.com/problemset/problem/1553/B"),
    (6,  "Maximum GCD",                   "Beginner", 800,  "https://codeforces.com/problemset/problem/1370/A"),
    (7,  "GCD and LCM",                   "Easy",     1400, "https://codeforces.com/problemset/problem/1458/C"),
    (8,  "Count Pairs",                   "Medium",   1700, "https://codeforces.com/problemset/problem/1156/D"),
    (9,  "Primes and Composites",         "Easy",     1200, "https://codeforces.com/problemset/problem/1366/D"),
    (10, "Sieve of Eratosthenes (use)",   "Medium",   1600, "https://codeforces.com/problemset/problem/1211/E"),
    (11, "Sum Queries",                   "Hard",     2000, "https://codeforces.com/problemset/problem/1217/E"),
    (12, "p-Binary",                      "Easy",     1400, "https://codeforces.com/problemset/problem/1225/C"),
    (13, "Factorial Divisibility",        "Easy",     1400, "https://codeforces.com/problemset/problem/1538/D"),
    (14, "Choose Two Distinct",           "Medium",   1500, "https://codeforces.com/problemset/problem/1538/C"),
    (15, "Ehab Fails",                    "Easy",     1300, "https://codeforces.com/problemset/problem/1239/B"),
]

# ── 8. Bit Manipulation ───────────────────────────────────────────────────────
topics["Bit Manipulation"] = [
    (1,  "Ternary XOR",                   "Easy",     1400, "https://codeforces.com/problemset/problem/1328/C"),
    (2,  "Magic Grid",                    "Medium",   1700, "https://codeforces.com/problemset/problem/1208/C"),
    (3,  "Adding Powers of 2",            "Easy",     1400, "https://codeforces.com/problemset/problem/1217/C"),
    (4,  "AND Segment",                   "Hard",     2000, "https://codeforces.com/problemset/problem/1370/F"),
    (5,  "XOR-ranges",                    "Medium",   1800, "https://codeforces.com/problemset/problem/1548/E"),
    (6,  "Number of Pairs",               "Medium",   1700, "https://codeforces.com/problemset/problem/1451/E"),
    (7,  "Ehab's Last Theorem (XOR)",     "Hard",     2000, "https://codeforces.com/problemset/problem/1364/D"),
    (8,  "Sum XOR Product",               "Medium",   1700, "https://codeforces.com/problemset/problem/1628/D"),
    (9,  "Path XOR",                      "Hard",     2000, "https://codeforces.com/problemset/problem/1325/C"),
    (10, "XOR-tree",                      "Hard",     2200, "https://codeforces.com/problemset/problem/1516/E"),
    (11, "Bits and Pieces",               "Hard",     2100, "https://codeforces.com/problemset/problem/1208/F"),
    (12, "Binary XOR",                    "Easy",     1300, "https://codeforces.com/problemset/problem/1554/A"),
]

# ── 9. Stack & Monotonic Stack ────────────────────────────────────────────────
topics["Stack & Monotonic Stack"] = [
    (1,  "Make It Good",                  "Easy",     1200, "https://codeforces.com/problemset/problem/1385/C"),
    (2,  "Valid BFS",                     "Medium",   1700, "https://codeforces.com/problemset/problem/1037/D"),
    (3,  "Balanced Brackets",             "Easy",     1300, "https://codeforces.com/problemset/problem/1452/C"),
    (4,  "Next Greater Element",          "Medium",   1600, "https://codeforces.com/problemset/problem/1313/C2"),
    (5,  "Restore the Queue",             "Medium",   1700, "https://codeforces.com/problemset/problem/1506/E"),
    (6,  "Array Stabilization",           "Medium",   1600, "https://codeforces.com/problemset/problem/1616/D"),
    (7,  "Periodic Sequence",             "Medium",   1600, "https://codeforces.com/problemset/problem/1364/C"),
    (8,  "Rectangle Queries",             "Medium",   1800, "https://codeforces.com/problemset/problem/1093/D"),
    (9,  "Distinct Values",               "Medium",   1700, "https://codeforces.com/problemset/problem/1156/E"),
    (10, "Monotone Queue Sliding Min",    "Hard",     2000, "https://codeforces.com/problemset/problem/1195/E"),
    (11, "Sereja and Suffixes",           "Medium",   1600, "https://codeforces.com/problemset/problem/381/C"),
    (12, "Queue of Students",             "Easy",     1300, "https://codeforces.com/problemset/problem/1529/B"),
]

# ── 10. Graphs — BFS ──────────────────────────────────────────────────────────
topics["Graphs — BFS"] = [
    (1,  "Learning Languages",            "Easy",     1400, "https://codeforces.com/problemset/problem/277/A"),
    (2,  "Messenger",                     "Medium",   1600, "https://codeforces.com/problemset/problem/1060/E"),
    (3,  "Shortest Path (01 BFS)",        "Medium",   1700, "https://codeforces.com/problemset/problem/590/C"),
    (4,  "Tree Painting",                 "Hard",     2000, "https://codeforces.com/problemset/problem/1187/E"),
    (5,  "Complete Tripartite",           "Medium",   1700, "https://codeforces.com/problemset/problem/1228/D"),
    (6,  "Watching Fireworks is Fun",     "Medium",   1800, "https://codeforces.com/problemset/problem/372/E"),
    (7,  "Bipartite Reachability",        "Medium",   1700, "https://codeforces.com/problemset/problem/1234/E"),
    (8,  "Minimum Spanning Graph",        "Easy",     1400, "https://codeforces.com/problemset/problem/1513/D"),
    (9,  "BFS on Grid with Obstacles",    "Medium",   1600, "https://codeforces.com/problemset/problem/1365/E"),
    (10, "Particle Collider",             "Medium",   1800, "https://codeforces.com/problemset/problem/1486/E"),
    (11, "Roads not only in Berland",     "Medium",   1700, "https://codeforces.com/problemset/problem/196/E"),
    (12, "Graph Connectivity",            "Easy",     1200, "https://codeforces.com/problemset/problem/1557/C"),
]

# ── 11. Graphs — DFS & Trees ──────────────────────────────────────────────────
topics["Graphs — DFS & Trees"] = [
    (1,  "Three Paths on a Tree",         "Medium",   1700, "https://codeforces.com/problemset/problem/1294/F"),
    (2,  "Vasya and a Tree",              "Hard",     2100, "https://codeforces.com/problemset/problem/1076/E"),
    (3,  "Lomsat Gelral (DSU on tree)",   "Hard",     2200, "https://codeforces.com/problemset/problem/600/E"),
    (4,  "Journey",                       "Medium",   1700, "https://codeforces.com/problemset/problem/839/C"),
    (5,  "Linova and Kingdom",            "Medium",   1700, "https://codeforces.com/problemset/problem/1337/C"),
    (6,  "Directing Edges (Topo Sort)",   "Medium",   1800, "https://codeforces.com/problemset/problem/1385/E"),
    (7,  "Subtree Queries",               "Hard",     2000, "https://codeforces.com/problemset/problem/1527/G"),
    (8,  "Bridges in Graph (Tarjan)",     "Hard",     2000, "https://codeforces.com/problemset/problem/1517/E"),
    (9,  "LCA (Binary Lifting)",          "Medium",   1800, "https://codeforces.com/problemset/problem/1304/E"),
    (10, "Tree DP — Rerooting",           "Hard",     2000, "https://codeforces.com/problemset/problem/1187/E"),
    (11, "Heavy-Light Decomposition",     "Expert",   2400, "https://codeforces.com/problemset/problem/342/E"),
    (12, "Centroid Decomposition",        "Expert",   2500, "https://codeforces.com/problemset/problem/321/C"),
    (13, "Distance in Tree",              "Medium",   1700, "https://codeforces.com/problemset/problem/161/D"),
    (14, "Euler Tour on Tree",            "Medium",   1800, "https://codeforces.com/problemset/problem/1209/H"),
]

# ── 12. Shortest Path ─────────────────────────────────────────────────────────
topics["Shortest Path"] = [
    (1,  "Dijkstra on Grid",              "Medium",   1700, "https://codeforces.com/problemset/problem/1514/E"),
    (2,  "Dijkstra + Segment Tree Graph", "Hard",     2000, "https://codeforces.com/problemset/problem/787/D"),
    (3,  "Bellman-Ford Negative Cycle",   "Medium",   1700, "https://codeforces.com/problemset/problem/1514/D"),
    (4,  "Floyd-Warshall Applications",   "Medium",   1700, "https://codeforces.com/problemset/problem/1486/D"),
    (5,  "01-BFS Shortest Path",          "Medium",   1700, "https://codeforces.com/problemset/problem/590/C"),
    (6,  "Cheapest Shortest Path",        "Medium",   1800, "https://codeforces.com/problemset/problem/1486/E"),
    (7,  "Find the City",                 "Medium",   1600, "https://codeforces.com/problemset/problem/1257/F"),
    (8,  "Network Delay (Dijkstra)",      "Medium",   1600, "https://codeforces.com/problemset/problem/1513/E"),
    (9,  "Road Reconstruction",           "Hard",     2100, "https://codeforces.com/problemset/problem/1239/D"),
    (10, "SSSP with Updates",             "Hard",     2200, "https://codeforces.com/problemset/problem/1163/F"),
]

# ── 13. Dynamic Programming — 1D ──────────────────────────────────────────────
topics["DP — 1D"] = [
    (1,  "Celex Update",                  "Medium",   1500, "https://codeforces.com/problemset/problem/1358/C"),
    (2,  "Move Brackets",                 "Easy",     1100, "https://codeforces.com/problemset/problem/1374/C"),
    (3,  "Replace by MEX",                "Medium",   1600, "https://codeforces.com/problemset/problem/1375/D"),
    (4,  "Linova and Kingdom",            "Medium",   1700, "https://codeforces.com/problemset/problem/1337/C"),
    (5,  "Petya and Strings",             "Easy",     1100, "https://codeforces.com/problemset/problem/1038B/B"),
    (6,  "Beautiful Array (DP Knapsack)", "Hard",     1900, "https://codeforces.com/problemset/problem/1155/D"),
    (7,  "Prefix Sum DP",                 "Medium",   1600, "https://codeforces.com/problemset/problem/1285/E"),
    (8,  "Maximum Subarray Variant",      "Medium",   1700, "https://codeforces.com/problemset/problem/1527/D"),
    (9,  "House Robber Variant",          "Medium",   1600, "https://codeforces.com/problemset/problem/1418/E"),
    (10, "LIS (Patience Sort O(nlogn))",  "Hard",     2000, "https://codeforces.com/problemset/problem/1436/D"),
    (11, "Coin Change Variant",           "Medium",   1600, "https://codeforces.com/problemset/problem/1517/D"),
    (12, "K-periodic Garland",            "Hard",     2200, "https://codeforces.com/problemset/problem/1353/E"),
    (13, "Johnny and Another Drop",       "Medium",   1500, "https://codeforces.com/problemset/problem/1361/C"),
    (14, "New Year Parties",              "Medium",   1600, "https://codeforces.com/problemset/problem/1283/E"),
]

# ── 14. Dynamic Programming — 2D / Interval ───────────────────────────────────
topics["DP — 2D & Interval"] = [
    (1,  "Grid DP Basics",                "Medium",   1600, "https://codeforces.com/problemset/problem/1499/D"),
    (2,  "LCS Variant",                   "Medium",   1700, "https://codeforces.com/problemset/problem/463/D"),
    (3,  "Politics (Tree DP)",            "Medium",   1600, "https://codeforces.com/problemset/problem/1061/E"),
    (4,  "TV Show (Greedy DP)",           "Hard",     2000, "https://codeforces.com/problemset/problem/1061/D"),
    (5,  "Interval DP (Merge Stones)",    "Hard",     2000, "https://codeforces.com/problemset/problem/1509/F"),
    (6,  "Palindrome DP",                 "Hard",     2000, "https://codeforces.com/problemset/problem/1527/F"),
    (7,  "Edit Distance Variant",         "Medium",   1800, "https://codeforces.com/problemset/problem/154/E"),
    (8,  "DP on Subsets (Bitmask)",       "Hard",     2000, "https://codeforces.com/problemset/problem/1209/E"),
    (9,  "Matrix Chain Multiplication",   "Hard",     2000, "https://codeforces.com/problemset/problem/1108/E"),
    (10, "Partition DP",                  "Hard",     2100, "https://codeforces.com/problemset/problem/1556/H"),
    (11, "DP + Binary Search (SMAWK)",    "Expert",   2500, "https://codeforces.com/problemset/problem/1083/E"),
    (12, "Digit DP",                      "Hard",     2000, "https://codeforces.com/problemset/problem/1529/E"),
    (13, "Square Subsequences",           "Hard",     2200, "https://codeforces.com/problemset/problem/1535/F"),
]

# ── 15. Dynamic Programming — Bitmask ────────────────────────────────────────
topics["DP — Bitmask"] = [
    (1,  "Bitmask DP on Subsets",         "Hard",     2000, "https://codeforces.com/problemset/problem/1209/E"),
    (2,  "TSP Variant",                   "Hard",     2100, "https://codeforces.com/problemset/problem/1536/E"),
    (3,  "DP on Profiles",                "Expert",   2400, "https://codeforces.com/problemset/problem/1332/F"),
    (4,  "Assign Tasks (Bitmask)",        "Hard",     2000, "https://codeforces.com/problemset/problem/1209/G"),
    (5,  "SOS DP (Sum Over Subsets)",     "Hard",     2100, "https://codeforces.com/problemset/problem/383/E"),
    (6,  "Count Spanning Trees",          "Hard",     2200, "https://codeforces.com/problemset/problem/1163/E"),
    (7,  "Graph Coloring DP",             "Hard",     2000, "https://codeforces.com/problemset/problem/553/E"),
    (8,  "Matching Bitmask",              "Hard",     2000, "https://codeforces.com/problemset/problem/1484/E"),
]

# ── 16. Data Structures (Segment Tree / BIT / Sparse Table) ──────────────────
topics["Data Structures"] = [
    (1,  "Range Sum Query (BIT)",         "Medium",   1600, "https://codeforces.com/problemset/problem/629/D"),
    (2,  "Range Minimum Query",           "Medium",   1700, "https://codeforces.com/problemset/problem/1093/G"),
    (3,  "Segment Tree with Lazy",        "Hard",     2000, "https://codeforces.com/problemset/problem/1217/E"),
    (4,  "Segment Tree Beats",            "Expert",   2600, "https://codeforces.com/problemset/problem/1515/H"),
    (5,  "Persistent Segment Tree",       "Expert",   2500, "https://codeforces.com/problemset/problem/786/C"),
    (6,  "DSU (Union Find) on Tree",      "Hard",     2200, "https://codeforces.com/problemset/problem/600/E"),
    (7,  "Sparse Table RMQ",              "Medium",   1700, "https://codeforces.com/problemset/problem/1579/E"),
    (8,  "Treap / Policy Tree",           "Expert",   2400, "https://codeforces.com/problemset/problem/1478/D"),
    (9,  "Merge Sort Tree",               "Expert",   2300, "https://codeforces.com/problemset/problem/1093/G"),
    (10, "Monotonic Deque",               "Hard",     2000, "https://codeforces.com/problemset/problem/1195/E"),
    (11, "Fenwick Tree (BIT 2D)",         "Hard",     2100, "https://codeforces.com/problemset/problem/1290/E"),
    (12, "Square Root Decomposition",     "Hard",     2200, "https://codeforces.com/problemset/problem/940/F"),
    (13, "Euler Tour + BIT",              "Hard",     2000, "https://codeforces.com/problemset/problem/1076/E"),
    (14, "Interval Tree",                 "Hard",     2100, "https://codeforces.com/problemset/problem/1651/F"),
]

# ── 17. Combinatorics & Counting ──────────────────────────────────────────────
topics["Combinatorics"] = [
    (1,  "Count the Arrays",              "Medium",   1800, "https://codeforces.com/problemset/problem/1362/D"),
    (2,  "Binomial Coefficients",         "Medium",   1600, "https://codeforces.com/problemset/problem/1479/B"),
    (3,  "Stars and Bars",                "Easy",     1400, "https://codeforces.com/problemset/problem/1549/B"),
    (4,  "Inclusion-Exclusion",           "Hard",     2100, "https://codeforces.com/problemset/problem/900/E"),
    (5,  "Catalan Numbers",               "Medium",   1800, "https://codeforces.com/problemset/problem/1285/F"),
    (6,  "Derangements",                  "Hard",     2000, "https://codeforces.com/problemset/problem/1496/E"),
    (7,  "Modular Arithmetic (Fermat)",   "Medium",   1600, "https://codeforces.com/problemset/problem/1279/D"),
    (8,  "Permutation Counting",          "Hard",     2200, "https://codeforces.com/problemset/problem/1342/F"),
    (9,  "Burnside's Lemma",              "Expert",   2400, "https://codeforces.com/problemset/problem/1209/G"),
    (10, "Kirchhoff's Theorem",           "Expert",   2500, "https://codeforces.com/problemset/problem/1163/E"),
    (11, "Lucas Theorem",                 "Hard",     2000, "https://codeforces.com/problemset/problem/1034/E"),
    (12, "Generating Functions",          "Expert",   2600, "https://codeforces.com/problemset/problem/1264/F"),
]

# ── 18. Divide & Conquer ──────────────────────────────────────────────────────
topics["Divide & Conquer"] = [
    (1,  "Merge Sort Inversions",         "Medium",   1700, "https://codeforces.com/problemset/problem/597/C"),
    (2,  "CDQ Divide & Conquer",          "Hard",     2200, "https://codeforces.com/problemset/problem/1093/G"),
    (3,  "Closest Pair of Points",        "Hard",     2000, "https://codeforces.com/problemset/problem/distance/A"),
    (4,  "Power Tower",                   "Medium",   1700, "https://codeforces.com/problemset/problem/906/D"),
    (5,  "Divide on Segments",            "Medium",   1800, "https://codeforces.com/problemset/problem/1468/H"),
    (6,  "Offline D&C DP Optimization",   "Expert",   2500, "https://codeforces.com/problemset/problem/1083/E"),
    (7,  "Randomized D&C",                "Hard",     2000, "https://codeforces.com/problemset/problem/1536/F"),
    (8,  "K-th Order Statistic",          "Medium",   1700, "https://codeforces.com/problemset/problem/1349/C"),
]

# ── 19. Game Theory ───────────────────────────────────────────────────────────
topics["Game Theory"] = [
    (1,  "Nim Game",                      "Medium",   1600, "https://codeforces.com/problemset/problem/1527/B"),
    (2,  "Sprague–Grundy Theorem",        "Hard",     2000, "https://codeforces.com/problemset/problem/1538/F"),
    (3,  "Green Hackenbush",              "Hard",     2200, "https://codeforces.com/problemset/problem/850/C"),
    (4,  "Stone Game",                    "Medium",   1700, "https://codeforces.com/problemset/problem/1527/C"),
    (5,  "Losing vs Winning States",      "Medium",   1600, "https://codeforces.com/problemset/problem/768/D"),
    (6,  "Staircase Nim",                 "Medium",   1700, "https://codeforces.com/problemset/problem/850/C"),
    (7,  "Misère Nim",                    "Medium",   1600, "https://codeforces.com/problemset/problem/1537/E"),
    (8,  "Game on a Graph",               "Hard",     2000, "https://codeforces.com/problemset/problem/1408/E"),
]

# ── 20. Flow & Matching ───────────────────────────────────────────────────────
topics["Flow & Matching"] = [
    (1,  "Max Bipartite Matching",        "Hard",     2000, "https://codeforces.com/problemset/problem/1292/D"),
    (2,  "Min Cut Max Flow",              "Hard",     2100, "https://codeforces.com/problemset/problem/1543/E"),
    (3,  "Hall's Marriage Theorem",       "Hard",     2200, "https://codeforces.com/problemset/problem/1253/E"),
    (4,  "Assignment Problem",            "Expert",   2400, "https://codeforces.com/problemset/problem/1441/E"),
    (5,  "Min Cost Max Flow",             "Expert",   2500, "https://codeforces.com/problemset/problem/1534/G"),
    (6,  "Network Flow",                  "Hard",     2100, "https://codeforces.com/problemset/problem/1368/F"),
    (7,  "Dilworth's Theorem",            "Hard",     2200, "https://codeforces.com/problemset/problem/1499/H"),
]

# ══════════════════════════════════════════════════════════════════════════════
#  Google-Specific Story-Based Problems
# ══════════════════════════════════════════════════════════════════════════════
google_style = [
    # (no, name, topic, difficulty, cf_rating, url)
    (1,  "Watermelon",                           "Implementation",     "Beginner", 800,  "https://codeforces.com/problemset/problem/4/A"),
    (2,  "Theatre Square",                       "Math",               "Beginner", 800,  "https://codeforces.com/problemset/problem/1/A"),
    (3,  "Magic Grid",                           "Bit Manipulation",   "Medium",   1700, "https://codeforces.com/problemset/problem/1208/C"),
    (4,  "Compress Words",                       "Strings",            "Hard",     2000, "https://codeforces.com/problemset/problem/1200/E"),
    (5,  "Linova and Kingdom",                   "Greedy + Trees",     "Medium",   1700, "https://codeforces.com/problemset/problem/1337/C"),
    (6,  "Three Paths on a Tree",                "Trees",              "Medium",   1700, "https://codeforces.com/problemset/problem/1294/F"),
    (7,  "Count the Arrays",                     "Combinatorics",      "Medium",   1800, "https://codeforces.com/problemset/problem/1362/D"),
    (8,  "K-periodic Garland",                   "DP",                 "Hard",     2200, "https://codeforces.com/problemset/problem/1353/E"),
    (9,  "The Fair Nut and Rectangles",          "Greedy + DP",        "Hard",     2200, "https://codeforces.com/problemset/problem/1083/E"),
    (10, "Vasya and a Tree",                     "Trees + DFS",        "Hard",     2100, "https://codeforces.com/problemset/problem/1076/E"),
    (11, "Balanced Playlist",                    "Two Pointers",       "Medium",   1700, "https://codeforces.com/problemset/problem/1237/D"),
    (12, "Bridges in Graph",                     "Graphs",             "Hard",     2000, "https://codeforces.com/problemset/problem/1517/E"),
    (13, "Lomsat Gelral",                        "Trees",              "Hard",     2200, "https://codeforces.com/problemset/problem/600/E"),
    (14, "Bohemian Rhapsody (Dijkstra variant)", "Shortest Path",      "Hard",     2000, "https://codeforces.com/problemset/problem/787/D"),
    (15, "Prefix Flip",                          "Strings",            "Medium",   1700, "https://codeforces.com/problemset/problem/1382/C"),
    (16, "Segment Tree Beats",                   "Data Structures",    "Expert",   2600, "https://codeforces.com/problemset/problem/1515/H"),
    (17, "Row GCD",                              "Math",               "Medium",   1800, "https://codeforces.com/problemset/problem/1458/C"),
    (18, "Power Tower",                          "Divide & Conquer",   "Medium",   1700, "https://codeforces.com/problemset/problem/906/D"),
    (19, "New Year Parties",                     "DP + Greedy",        "Medium",   1600, "https://codeforces.com/problemset/problem/1283/E"),
    (20, "Beautiful Array",                      "DP",                 "Hard",     1900, "https://codeforces.com/problemset/problem/1155/D"),
    (21, "Perform the Combo",                    "Strings + Prefix",   "Easy",     1200, "https://codeforces.com/problemset/problem/1221/C"),
    (22, "Sum Queries",                          "Segment Tree",       "Hard",     2000, "https://codeforces.com/problemset/problem/1217/E"),
    (23, "Complete Tripartite",                  "Graphs",             "Medium",   1700, "https://codeforces.com/problemset/problem/1228/D"),
    (24, "Journey",                              "DFS + DP",           "Medium",   1700, "https://codeforces.com/problemset/problem/839/C"),
    (25, "Number of Pairs",                      "Bit Manipulation",   "Medium",   1700, "https://codeforces.com/problemset/problem/1451/E"),
]

# ══════════════════════════════════════════════════════════════════════════════
#  BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════

def add_topic_sheet_cf(wb, sheet_name, rows):
    ws = wb.create_sheet(title=safe(sheet_name))

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = sheet_name
    c.fill      = TITLE_FILL
    c.font      = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    hdrs   = ["#", "Problem Name", "Difficulty", "CF Rating", "Contest/Prob", "Link"]
    widths = [5, 50, 12, 12, 14, 48]
    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = COLHDR_FILL; cell.font = WHITE
        cell.alignment = CENTER; cell.border = BOX
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    DIFF_COLOR = {"Beginner":"2E7D32","Easy":"F57F17","Medium":"E65100","Hard":"B71C1C","Expert":"6A1B9A"}

    for r, (no, name, diff, rating, url) in enumerate(rows, start=3):
        fill = DIFF_FILL.get(diff, PatternFill())
        parts = url.replace("https://codeforces.com/problemset/problem/", "").rstrip("/").split("/")
        contest_label = f"{parts[0]}/{parts[1]}" if len(parts) >= 2 else ""

        ws.cell(row=r, column=1, value=no).alignment      = CENTER
        ws.cell(row=r, column=2, value=name).alignment    = LEFT
        diff_cell = ws.cell(row=r, column=3, value=diff)
        diff_cell.alignment = CENTER
        diff_cell.font      = Font(bold=True, size=10, color=DIFF_COLOR.get(diff, "000000"))
        ws.cell(row=r, column=4, value=rating).alignment          = CENTER
        ws.cell(row=r, column=5, value=contest_label).alignment   = CENTER

        lc = ws.cell(row=r, column=6, value="Open →")
        lc.hyperlink  = url
        lc.font       = LINK
        lc.alignment  = LEFT

        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.fill   = fill
            cell.border = BOX
            if col not in (3, 6):
                cell.font = NORMAL
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A3"
    return ws


# ── Google-specific sheet ─────────────────────────────────────────────────────
def add_google_sheet(wb, rows):
    ws = wb.create_sheet(title="Google-Style Problems")

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = "Google-Style Story-Based Problems (Codeforces)"
    c.fill      = PatternFill("solid", fgColor="1A237E")
    c.font      = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 30

    hdrs   = ["#", "Problem Name", "Topic", "Difficulty", "CF Rating", "Contest/Prob", "Link"]
    widths = [5, 42, 24, 12, 12, 14, 46]
    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="283593")
        cell.font = WHITE; cell.alignment = CENTER; cell.border = BOX
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    DIFF_COLOR = {"Beginner":"2E7D32","Easy":"F57F17","Medium":"E65100","Hard":"B71C1C","Expert":"6A1B9A"}
    ALT1 = PatternFill("solid", fgColor="E8EAF6")
    ALT2 = PatternFill("solid", fgColor="FFFFFF")

    for r, (no, name, topic, diff, rating, url) in enumerate(rows, start=3):
        fill  = ALT1 if r % 2 == 1 else ALT2
        parts = url.replace("https://codeforces.com/problemset/problem/", "").rstrip("/").split("/")
        cl    = f"{parts[0]}/{parts[1]}" if len(parts) >= 2 else ""

        ws.cell(row=r, column=1, value=no).alignment      = CENTER
        ws.cell(row=r, column=2, value=name).alignment    = LEFT
        ws.cell(row=r, column=3, value=topic).alignment   = LEFT
        dc = ws.cell(row=r, column=4, value=diff)
        dc.alignment = CENTER
        dc.font      = Font(bold=True, size=10, color=DIFF_COLOR.get(diff, "000000"))
        ws.cell(row=r, column=5, value=rating).alignment  = CENTER
        ws.cell(row=r, column=6, value=cl).alignment      = CENTER
        lc = ws.cell(row=r, column=7, value="Open →")
        lc.hyperlink = url; lc.font = LINK; lc.alignment = LEFT

        for col in range(1, 8):
            cell = ws.cell(row=r, column=col)
            cell.fill   = fill; cell.border = BOX
            if col not in (4, 7):
                cell.font = NORMAL
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A3"


# ── Build all topic sheets ────────────────────────────────────────────────────
for name, rows in topics.items():
    add_topic_sheet_cf(wb, name, rows)

add_google_sheet(wb, google_style)

# ── Overview sheet ────────────────────────────────────────────────────────────
ws_idx = wb.create_sheet(title="Overview", index=0)
ws_idx.merge_cells("A1:D1")
c = ws_idx["A1"]
c.value     = "Codeforces Interview Prep — Overview"
c.fill      = TITLE_FILL
c.font      = Font(color="FFFFFF", bold=True, size=14)
c.alignment = CENTER
ws_idx.row_dimensions[1].height = 32

ws_idx.column_dimensions["A"].width = 32
ws_idx.column_dimensions["B"].width = 14
ws_idx.column_dimensions["C"].width = 18
ws_idx.column_dimensions["D"].width = 44

for col, h in enumerate(["Sheet", "# Problems", "Difficulty Range", "Description"], 1):
    cell = ws_idx.cell(row=2, column=col, value=h)
    cell.fill = COLHDR_FILL; cell.font = WHITE
    cell.alignment = CENTER; cell.border = BOX
ws_idx.row_dimensions[2].height = 20

overview_rows = [
    ("Implementation",       topics["Implementation"],       "800–1400",  "Straightforward coding, simulation"),
    ("Greedy",               topics["Greedy"],               "900–1800",  "Locally optimal choices, interviews' bread & butter"),
    ("Sorting",              topics["Sorting"],              "800–1800",  "Comparison, custom sort, inversions"),
    ("Binary Search",        topics["Binary Search"],        "1200–2400", "Search on answer, parametric, rotated arrays"),
    ("Two Pointers",         topics["Two Pointers"],         "900–2200",  "Sliding window, shrink/expand window"),
    ("Strings",              topics["Strings"],              "1200–2000", "Pattern matching, hashing, KMP, Z-function"),
    ("Math & Number Theory", topics["Math & Number Theory"], "800–2000",  "GCD, LCM, primes, Euler, modular arithmetic"),
    ("Bit Manipulation",     topics["Bit Manipulation"],     "1300–2200", "XOR tricks, bitmask, subset enumeration"),
    ("Stack & Monotonic",    topics["Stack & Monotonic Stack"], "1200–2000", "Monotonic stack, deque, next greater element"),
    ("Graphs — BFS",         topics["Graphs — BFS"],         "1400–2400", "Level-order traversal, 0-1 BFS, multi-source BFS"),
    ("Graphs — DFS & Trees", topics["Graphs — DFS & Trees"], "1700–2500", "DFS, tree DP, rerooting, LCA, HLD, centroid"),
    ("Shortest Path",        topics["Shortest Path"],        "1600–2200", "Dijkstra, Bellman-Ford, Floyd-Warshall, SSSP"),
    ("DP — 1D",              topics["DP — 1D"],              "1100–2200", "Linear DP, LIS, knapsack, prefix sum DP"),
    ("DP — 2D & Interval",   topics["DP — 2D & Interval"],  "1600–2500", "Grid DP, interval DP, LCS, bitmask DP"),
    ("DP — Bitmask",         topics["DP — Bitmask"],         "2000–2600", "TSP, SOS DP, DP on subsets"),
    ("Data Structures",      topics["Data Structures"],      "1600–2600", "Segment tree, BIT, sparse table, DSU, SQRT decomp"),
    ("Combinatorics",        topics["Combinatorics"],        "1400–2600", "Counting, nCr, inclusion-exclusion, generating functions"),
    ("Divide & Conquer",     topics["Divide & Conquer"],     "1700–2500", "Merge sort, CDQ, D&C DP optimization"),
    ("Game Theory",          topics["Game Theory"],          "1600–2200", "Nim, Sprague-Grundy, Grundy values"),
    ("Flow & Matching",      topics["Flow & Matching"],      "2000–2500", "Max flow, bipartite matching, min cost flow"),
    ("Google-Style Problems","*see sheet*",                  "800–2600",  "Story-based problems commonly seen in Google-style rounds"),
]

TOPIC_F  = PatternFill("solid", fgColor="FBE9E7")
GOOGLE_F = PatternFill("solid", fgColor="E8EAF6")

for r_idx, row in enumerate(overview_rows, start=3):
    name_val = row[0]
    count_val = len(row[1]) if isinstance(row[1], list) else row[1]
    diff_range = row[2]
    desc = row[3]

    fill = GOOGLE_F if name_val.startswith("Google") else TOPIC_F

    ws_idx.cell(row=r_idx, column=1, value=name_val).fill   = fill
    ws_idx.cell(row=r_idx, column=1).font      = BOLD
    ws_idx.cell(row=r_idx, column=1).alignment = LEFT
    ws_idx.cell(row=r_idx, column=1).border    = BOX

    ws_idx.cell(row=r_idx, column=2, value=count_val).fill   = fill
    ws_idx.cell(row=r_idx, column=2).font      = NORMAL
    ws_idx.cell(row=r_idx, column=2).alignment = CENTER
    ws_idx.cell(row=r_idx, column=2).border    = BOX

    ws_idx.cell(row=r_idx, column=3, value=diff_range).fill   = fill
    ws_idx.cell(row=r_idx, column=3).font      = NORMAL
    ws_idx.cell(row=r_idx, column=3).alignment = CENTER
    ws_idx.cell(row=r_idx, column=3).border    = BOX

    ws_idx.cell(row=r_idx, column=4, value=desc).fill   = fill
    ws_idx.cell(row=r_idx, column=4).font      = NORMAL
    ws_idx.cell(row=r_idx, column=4).alignment = LEFT
    ws_idx.cell(row=r_idx, column=4).border    = BOX
    ws_idx.row_dimensions[r_idx].height = 20

# ── Legend row ────────────────────────────────────────────────────────────────
legend_row = r_idx + 2
ws_idx.merge_cells(f"A{legend_row}:D{legend_row}")
lc = ws_idx.cell(row=legend_row, column=1, value="Difficulty Legend")
lc.fill = OVERVIEW_HDR; lc.font = WHITE; lc.alignment = CENTER; lc.border = BOX
ws_idx.row_dimensions[legend_row].height = 24

legend_data = [
    ("Beginner  (800–1000)", "2E7D32", "LeetCode Easy equivalent — warm-up"),
    ("Easy      (1100–1400)", "F57F17", "LeetCode Easy-Medium — core interview problems"),
    ("Medium    (1500–1800)", "E65100", "LeetCode Medium-Hard — most interview targets"),
    ("Hard      (1900–2200)", "B71C1C", "LeetCode Hard — Google / competitive rounds"),
    ("Expert    (2300+)",     "6A1B9A", "Competitive programming — optional stretch goals"),
]
for i, (label, color, desc) in enumerate(legend_data):
    r = legend_row + 1 + i
    ws_idx.merge_cells(f"A{r}:B{r}")
    cell = ws_idx.cell(row=r, column=1, value=label)
    cell.fill      = DIFF_FILL[label.split()[0]]
    cell.font      = Font(bold=True, color=color, size=10)
    cell.alignment = LEFT; cell.border = BOX

    ws_idx.merge_cells(f"C{r}:D{r}")
    dc = ws_idx.cell(row=r, column=3, value=desc)
    dc.fill      = DIFF_FILL[label.split()[0]]
    dc.font      = NORMAL
    dc.alignment = LEFT; dc.border = BOX
    ws_idx.row_dimensions[r].height = 18

ws_idx.freeze_panes = "A3"

# ── Resources sheet ───────────────────────────────────────────────────────────
ws_res = wb.create_sheet(title="Resources & Plan")

ws_res.merge_cells("A1:C1")
c = ws_res["A1"]
c.value     = "Codeforces Prep — Resources & Study Tips"
c.fill      = PLAN_FILL
c.font      = Font(color="FFFFFF", bold=True, size=14)
c.alignment = CENTER
ws_res.row_dimensions[1].height = 30

ws_res.column_dimensions["A"].width = 30
ws_res.column_dimensions["B"].width = 55
ws_res.column_dimensions["C"].width = 35

for col, h in enumerate(["Resource", "Link", "Notes"], 1):
    cell = ws_res.cell(row=2, column=col, value=h)
    cell.fill = PatternFill("solid", fgColor="2E7D32")
    cell.font = WHITE; cell.alignment = CENTER; cell.border = BOX
ws_res.row_dimensions[2].height = 20

resources = [
    ("Codeforces Problemset",        "https://codeforces.com/problemset",                          "Filter by tag + rating"),
    ("A2OJ Ladder",                   "https://a2oj.netlify.app/",                                 "Curated ladders by rating"),
    ("CSES Problem Set",              "https://cses.fi/problemset/",                               "Best structured CP problem set"),
    ("CP-Algorithms",                 "https://cp-algorithms.com/",                                "Algorithm explanations with proofs"),
    ("Codeforces EDU",                "https://codeforces.com/edu/courses",                        "Official interactive courses"),
    ("AtCoder Problems",              "https://kenkoooo.com/atcoder/",                             "Companion to CF, similar style"),
    ("Google Kick Start Archive",     "https://codingcompetitions.withgoogle.com/kickstart",       "Google's own story-based rounds"),
    ("Codeforces Rating Calculator",  "https://codeforces.com/ratings",                           "Track and plan your growth"),
    ("Competitive Programmer's Handbook", "https://cses.fi/book/book.pdf",                        "Free book — covers everything"),
    ("NeetCode (LeetCode bridge)",    "https://neetcode.io/roadmap",                              "Use alongside CF for interviews"),
]

ALT1 = PatternFill("solid", fgColor="E8F5E9")
ALT2 = PatternFill("solid", fgColor="FFFFFF")
for r_idx, (name, url, note) in enumerate(resources, start=3):
    fill = ALT1 if r_idx % 2 == 1 else ALT2
    ws_res.cell(row=r_idx, column=1, value=name).fill   = fill
    ws_res.cell(row=r_idx, column=1).font      = BOLD
    ws_res.cell(row=r_idx, column=1).alignment = LEFT
    ws_res.cell(row=r_idx, column=1).border    = BOX

    lc = ws_res.cell(row=r_idx, column=2, value=url)
    lc.hyperlink = url; lc.font = LINK; lc.fill = fill
    lc.alignment = LEFT; lc.border = BOX

    ws_res.cell(row=r_idx, column=3, value=note).fill   = fill
    ws_res.cell(row=r_idx, column=3).font      = NORMAL
    ws_res.cell(row=r_idx, column=3).alignment = LEFT
    ws_res.cell(row=r_idx, column=3).border    = BOX
    ws_res.row_dimensions[r_idx].height = 18

# ── Phase plan ────────────────────────────────────────────────────────────────
phase_start = r_idx + 2
ws_res.merge_cells(f"A{phase_start}:C{phase_start}")
ph = ws_res.cell(row=phase_start, column=1, value="Suggested CF Preparation Phases")
ph.fill = PLAN_FILL; ph.font = Font(color="FFFFFF", bold=True, size=12)
ph.alignment = CENTER; ph.border = BOX
ws_res.row_dimensions[phase_start].height = 26

phase_hdrs = ["Phase", "Topics", "Target Rating Range"]
for col, h in enumerate(phase_hdrs, 1):
    cell = ws_res.cell(row=phase_start + 1, column=col, value=h)
    cell.fill = PatternFill("solid", fgColor="2E7D32")
    cell.font = WHITE; cell.alignment = CENTER; cell.border = BOX

phases = [
    ("Phase 1 — Foundation\n(Weeks 1–2)",  "Implementation, Greedy, Sorting, Basic Math",                 "800–1200"),
    ("Phase 2 — Core Skills\n(Weeks 3–5)", "Binary Search, Two Pointers, Strings, Graphs BFS/DFS",        "1200–1600"),
    ("Phase 3 — Interview Level\n(Wk 6–9)","DP 1D & 2D, Trees, Shortest Path, Stack & Queue",            "1600–2000"),
    ("Phase 4 — Advanced\n(Wk 10–12)",     "Data Structures, Combinatorics, Bitmask DP, Game Theory",    "2000–2400"),
    ("Phase 5 — Contest Practice",          "Participate in Div. 2 / Div. 3 rounds weekly",               "2400+"),
]
PHASE_COLORS = ["E3F2FD", "E8F5E9", "FFF8E1", "FCE4EC", "F3E5F5"]
for i, (ph_name, topics_str, rating) in enumerate(phases):
    r = phase_start + 2 + i
    fill = PatternFill("solid", fgColor=PHASE_COLORS[i])
    ws_res.cell(row=r, column=1, value=ph_name).fill   = fill
    ws_res.cell(row=r, column=1).font      = BOLD
    ws_res.cell(row=r, column=1).alignment = LEFT
    ws_res.cell(row=r, column=1).border    = BOX
    ws_res.cell(row=r, column=2, value=topics_str).fill   = fill
    ws_res.cell(row=r, column=2).font      = NORMAL
    ws_res.cell(row=r, column=2).alignment = LEFT
    ws_res.cell(row=r, column=2).border    = BOX
    ws_res.cell(row=r, column=3, value=rating).fill   = fill
    ws_res.cell(row=r, column=3).font      = NORMAL
    ws_res.cell(row=r, column=3).alignment = CENTER
    ws_res.cell(row=r, column=3).border    = BOX
    ws_res.row_dimensions[r].height = 34

ws_res.freeze_panes = "A3"

# ── Save ──────────────────────────────────────────────────────────────────────
out = r"c:\Users\MaitySom\Desktop\Somneel\Somneel_Maity\K_Tech\DSA\Codeforces_Interview_Prep.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Sheets: {len(wb.sheetnames)} — {wb.sheetnames}")
